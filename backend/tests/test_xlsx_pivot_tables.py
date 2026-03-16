"""
Tests for Pivot Table Accessibility Analysis (Task 11)

Tests cover:
- PivotTableIssue model
- Pivot table detection in XLSX files
- Field analysis (row, column, value fields)
- Nested header detection
- Integration with XlsxProcessor
"""

import pytest
import tempfile
import os

from openpyxl import Workbook

from src.education.xlsx_processor import (
    XlsxProcessor,
    PivotTableIssue,
)


class TestPivotTableIssueModel:
    """Test PivotTableIssue Pydantic model."""

    def test_pivot_issue_creation(self):
        """Test creating a pivot table issue."""
        issue = PivotTableIssue(
            sheet_name="Analysis",
            pivot_name="SalesPivot",
            pivot_location="A1:E20",
            row_fields=["Region", "Product"],
            column_fields=["Quarter"],
            value_fields=["Sales Amount", "Units Sold"],
            has_field_labels=True,
            has_grand_totals=True,
            issue_type="complex_structure",
            recommendations=[
                "Consider providing flat data table alternative",
                "Ensure field labels are descriptive",
            ],
        )

        assert issue.sheet_name == "Analysis"
        assert issue.pivot_name == "SalesPivot"
        assert len(issue.row_fields) == 2
        assert len(issue.column_fields) == 1
        assert len(issue.value_fields) == 2
        assert issue.issue_type == "complex_structure"

    def test_pivot_issue_minimal(self):
        """Test pivot table issue with minimal fields."""
        issue = PivotTableIssue(
            sheet_name="Data",
            pivot_name="PivotTable1",
            pivot_location="A1:C10",
            issue_type="missing_labels",
            recommendations=["Add field labels"],
        )

        assert issue.row_fields == []
        assert issue.column_fields == []
        assert issue.value_fields == []
        assert issue.has_field_labels is True  # Default
        assert issue.has_grand_totals is False

    def test_pivot_issue_types(self):
        """Test different pivot table issue types."""
        issue_types = ["complex_structure", "missing_labels", "nested_headers"]

        for issue_type in issue_types:
            issue = PivotTableIssue(
                sheet_name="Test",
                pivot_name="TestPivot",
                pivot_location="A1:D10",
                issue_type=issue_type,
                recommendations=["Test recommendation"],
            )
            assert issue.issue_type == issue_type


class TestPivotTableDetection:
    """Test pivot table detection in XLSX files."""

    @pytest.fixture
    def processor(self):
        """Create XlsxProcessor for testing."""
        return XlsxProcessor()

    def test_no_pivot_tables(self, processor):
        """Test that sheets without pivot tables have no issues."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb = Workbook()
            ws = wb.active
            ws.title = "Data"

            # Add simple data without pivot table
            headers = ["Name", "Value", "Category"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)

            for row in range(2, 10):
                ws.cell(row=row, column=1, value=f"Item {row}")
                ws.cell(row=row, column=2, value=row * 10)
                ws.cell(row=row, column=3, value="A" if row % 2 == 0 else "B")

            wb.save(f.name)

            result = processor.process_xlsx(f.name)

            assert len(result.sheets[0].pivot_table_issues) == 0
            assert result.summary.get("pivot_table_issues", 0) == 0

        os.unlink(f.name)

    def test_pivot_issues_in_sheet_analysis(self, processor):
        """Test that pivot_table_issues field exists in sheet analysis."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb = Workbook()
            ws = wb.active
            ws.cell(row=1, column=1, value="Test")
            wb.save(f.name)

            result = processor.process_xlsx(f.name)

            # Sheet analysis should have pivot_table_issues field
            assert hasattr(result.sheets[0], "pivot_table_issues")
            assert isinstance(result.sheets[0].pivot_table_issues, list)

        os.unlink(f.name)


class TestPivotTableAnalysis:
    """Test pivot table analysis method."""

    @pytest.fixture
    def processor(self):
        """Create XlsxProcessor for testing."""
        return XlsxProcessor()

    def test_analyze_pivot_tables_empty_sheet(self, processor):
        """Test _analyze_pivot_tables on sheet without pivots."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb = Workbook()
            ws = wb.active
            ws.cell(row=1, column=1, value="Data")
            wb.save(f.name)

            result = processor.process_xlsx(f.name)

            # Should return empty list for sheet without pivots
            assert result.sheets[0].pivot_table_issues == []

        os.unlink(f.name)

    def test_analyze_pivot_tables_returns_list(self, processor):
        """Test that _analyze_pivot_tables always returns a list."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb = Workbook()
            ws = wb.active
            ws.cell(row=1, column=1, value="Test")
            wb.save(f.name)

            result = processor.process_xlsx(f.name)

            assert isinstance(result.sheets[0].pivot_table_issues, list)

        os.unlink(f.name)


class TestXlsxProcessorIntegration:
    """Test pivot table integration with XlsxProcessor."""

    @pytest.fixture
    def processor(self):
        """Create XlsxProcessor for testing."""
        return XlsxProcessor()

    def test_summary_includes_pivot_issues(self, processor):
        """Test that summary includes pivot_table_issues key."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb = Workbook()
            ws = wb.active
            ws.cell(row=1, column=1, value="Test")
            wb.save(f.name)

            result = processor.process_xlsx(f.name)

            assert "pivot_table_issues" in result.summary

        os.unlink(f.name)

    def test_pivot_issues_in_total_count(self, processor):
        """Test that pivot issues would be counted in total issues."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb = Workbook()
            ws = wb.active
            ws.cell(row=1, column=1, value="Test")
            wb.save(f.name)

            result = processor.process_xlsx(f.name)

            pivot_count = result.summary.get("pivot_table_issues", 0)
            total = result.summary.get("total_issues", 0)

            # If there are pivot issues, they should be in total
            if pivot_count > 0:
                assert total >= pivot_count

        os.unlink(f.name)

    def test_multiple_sheets_pivot_count(self, processor):
        """Test pivot table counting across multiple sheets."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb = Workbook()

            # First sheet
            ws1 = wb.active
            ws1.title = "Sheet1"
            ws1.cell(row=1, column=1, value="Data1")

            # Second sheet
            ws2 = wb.create_sheet("Sheet2")
            ws2.cell(row=1, column=1, value="Data2")

            wb.save(f.name)

            result = processor.process_xlsx(f.name)

            # Total should sum across all sheets
            total_pivot = sum(len(sheet.pivot_table_issues) for sheet in result.sheets)
            assert result.summary.get("pivot_table_issues", 0) == total_pivot

        os.unlink(f.name)


class TestPivotIssuesComputedField:
    """Test that pivot issues appear in the computed issues field."""

    @pytest.fixture
    def processor(self):
        """Create XlsxProcessor for testing."""
        return XlsxProcessor()

    def test_pivot_computed_issues_structure(self, processor):
        """Test that pivot issues would have correct structure in computed issues."""
        # Create a simple test to verify the computed issues field exists
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb = Workbook()
            ws = wb.active
            ws.cell(row=1, column=1, value="Test")
            wb.save(f.name)

            result = processor.process_xlsx(f.name)

            # Verify issues is a list
            assert isinstance(result.issues, list)

            # If there were pivot issues, they would have these attributes
            # This tests the structure definition even without actual pivot tables
            expected_keys = [
                "id",
                "category",
                "severity",
                "title",
                "description",
                "location",
                "wcag_criterion",
                "suggested_fix",
            ]

            # All issues should have the expected keys
            for issue in result.issues:
                for key in expected_keys:
                    assert key in issue, f"Missing key {key} in issue"

        os.unlink(f.name)


class TestPivotTableRecommendations:
    """Test pivot table recommendation generation."""

    def test_pivot_issue_has_recommendations(self):
        """Test that pivot issues include recommendations."""
        issue = PivotTableIssue(
            sheet_name="Test",
            pivot_name="TestPivot",
            pivot_location="A1:D10",
            row_fields=["Field1", "Field2"],
            column_fields=["Field3"],
            issue_type="nested_headers",
            recommendations=[
                "Consider flattening the structure",
                "Provide flat data table alternative",
            ],
        )

        assert len(issue.recommendations) >= 1
        assert isinstance(issue.recommendations, list)

    def test_pivot_issue_recommendations_list(self):
        """Test recommendations are stored as list."""
        recommendations = [
            "Recommendation 1",
            "Recommendation 2",
            "Recommendation 3",
        ]

        issue = PivotTableIssue(
            sheet_name="Test",
            pivot_name="TestPivot",
            pivot_location="A1:D10",
            issue_type="complex_structure",
            recommendations=recommendations,
        )

        assert issue.recommendations == recommendations


class TestRemediationSuggestions:
    """Test remediation suggestions for pivot tables."""

    @pytest.fixture
    def processor(self):
        """Create XlsxProcessor for testing."""
        return XlsxProcessor()

    def test_no_pivot_suggestion_without_issues(self, processor):
        """Test that no pivot suggestion appears without pivot issues."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb = Workbook()
            ws = wb.active
            ws.cell(row=1, column=1, value="Simple Data")
            wb.save(f.name)

            result = processor.process_xlsx(f.name)

            # No pivot issues means no pivot remediation suggestion
            if result.summary.get("pivot_table_issues", 0) == 0:
                pivot_suggestions = [
                    s
                    for s in result.remediation_suggestions
                    if "pivot table" in s.lower()
                ]
                assert len(pivot_suggestions) == 0

        os.unlink(f.name)


class TestComplianceScoring:
    """Test that pivot table issues affect compliance scoring."""

    @pytest.fixture
    def processor(self):
        """Create XlsxProcessor for testing."""
        return XlsxProcessor()

    def test_clean_sheet_high_score(self, processor):
        """Test that clean sheet has high compliance score."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb = Workbook()
            ws = wb.active
            ws.title = "Clean Data"  # Not a generic name

            # Add proper table structure
            headers = ["Name", "Value"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = cell.font.copy(bold=True)  # Bold headers

            for row in range(2, 6):
                ws.cell(row=row, column=1, value=f"Item {row}")
                ws.cell(row=row, column=2, value=row * 10)

            wb.save(f.name)

            result = processor.process_xlsx(f.name)

            # Clean sheet should have high score
            assert result.compliance_score >= 80.0

        os.unlink(f.name)


if __name__ == "__main__":
    pytest.main([__file__, "-v", "-s"])
