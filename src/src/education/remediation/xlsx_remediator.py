"""
Excel Remediator for Aelira Auto-Remediation Engine.

This module provides automatic remediation for accessibility issues in
Microsoft Excel spreadsheets (.xlsx files).

Supported auto-fixes:
- Rename generic sheet names
- Define table headers
- Add chart alt text
- Remove color-only indicators
- Add cell descriptions
- Freeze header rows for navigation
"""

import re
import logging
from typing import Any, Dict, List, Optional
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo

from .base import (
    BaseRemediator,
    RemediationIssue,
    IssueCategory,
    IssueSeverity,
    RemediationConfig,
)

logger = logging.getLogger(__name__)


class XlsxRemediator(BaseRemediator):
    """
    Remediator for Microsoft Excel spreadsheets (.xlsx).

    Automatically fixes accessibility issues including:
    - Generic sheet names (Sheet1, Sheet2)
    - Missing table headers
    - Charts without alt text
    - Color-only information indicators
    - Navigation issues (frozen panes)

    Usage:
        issues = [{'type': 'sheet', 'severity': 'high', ...}]
        remediator = XlsxRemediator('spreadsheet.xlsx', issues)
        result = remediator.remediate()
    """

    DOCUMENT_TYPE = "excel"
    SUPPORTED_EXTENSIONS = [".xlsx"]

    AUTO_FIXABLE_CATEGORIES = [
        IssueCategory.SHEET,
        IssueCategory.TABLE,
        IssueCategory.CHART,
        IssueCategory.COLOR,
        IssueCategory.NAVIGATION,
    ]

    # Generic sheet name patterns
    GENERIC_NAMES = [
        "sheet1",
        "sheet2",
        "sheet3",
        "sheet",
        "data",
        "new sheet",
        "worksheet",
    ]

    def __init__(
        self,
        file_path: str,
        issues: List[Dict[str, Any]],
        config: Optional[RemediationConfig] = None,
        ai_client: Optional[Any] = None,
    ):
        """Initialize the Excel remediator."""
        super().__init__(file_path, issues, config, ai_client)
        self._workbook: Optional[Any] = None

    def _load_document(self) -> Any:
        """Load the Excel workbook for editing."""
        logger.info(f"Loading Excel: {self.file_path}")
        self._workbook = load_workbook(self.file_path)
        return self._workbook

    def _save_document(self, document: Any) -> str:
        """Save the remediated Excel workbook."""
        output_path = self._get_output_path()
        logger.info(f"Saving remediated spreadsheet to: {output_path}")

        # Ensure output directory exists
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)

        document.save(output_path)
        return output_path

    def can_auto_fix(self, issue: RemediationIssue) -> bool:
        """
        Determine if an issue can be automatically fixed.

        Auto-fixable issues:
        - Sheet names: If we can generate meaningful name
        - Table headers: If first row exists
        - Chart alt text: With AI or placeholder
        - Color indicators: Add text alternatives
        - Navigation: Add frozen panes
        """
        if issue.category not in self.AUTO_FIXABLE_CATEGORIES:
            return False

        if issue.category == IssueCategory.SHEET:
            # Can rename if we have sheet data to infer name
            return True

        if issue.category == IssueCategory.TABLE:
            # Can add headers if data exists
            return bool(issue.metadata.get("has_data", True))

        if issue.category == IssueCategory.CHART:
            # Can add alt text with AI or placeholder
            return self.config.use_ai or True

        if issue.category == IssueCategory.COLOR:
            # Can add text indicators
            return True

        if issue.category == IssueCategory.NAVIGATION:
            # Can freeze panes
            return True

        return False

    def apply_fix(
        self, issue: RemediationIssue, document: Any, fix_content: str
    ) -> bool:
        """
        Apply a fix to the Excel workbook.

        Args:
            issue: The issue being fixed
            document: The Excel workbook object
            fix_content: The content to apply as the fix

        Returns:
            True if fix was applied successfully
        """
        try:
            if issue.category == IssueCategory.SHEET:
                return self._apply_sheet_name_fix(issue, document, fix_content)

            if issue.category == IssueCategory.TABLE:
                return self._apply_table_header_fix(issue, document, fix_content)

            if issue.category == IssueCategory.CHART:
                return self._apply_chart_fix(issue, document, fix_content)

            if issue.category == IssueCategory.COLOR:
                return self._apply_color_fix(issue, document, fix_content)

            if issue.category == IssueCategory.NAVIGATION:
                return self._apply_navigation_fix(issue, document, fix_content)

            return False

        except Exception as e:
            logger.error(f"Failed to apply fix for issue {issue.id}: {e}")
            return False

    def _apply_sheet_name_fix(
        self, issue: RemediationIssue, document: Any, new_name: str
    ) -> bool:
        """Rename a sheet to a more descriptive name."""
        try:
            sheet_name = issue.metadata.get("sheet_name", "")
            sheet_index = issue.metadata.get("sheet_index")

            if not sheet_name and sheet_index is not None:
                sheet_name = document.sheetnames[sheet_index]

            if sheet_name not in document.sheetnames:
                logger.warning(f"Sheet not found: {sheet_name}")
                return False

            worksheet = document[sheet_name]

            # Generate new name if not provided
            if not new_name or new_name == sheet_name:
                new_name = self._generate_sheet_name(worksheet, sheet_index)

            # Ensure unique name
            new_name = self._make_unique_name(document, new_name, sheet_name)

            # Rename the sheet
            worksheet.title = new_name

            logger.info(f"Renamed sheet '{sheet_name}' to '{new_name}'")
            return True

        except Exception as e:
            logger.error(f"Error renaming sheet: {e}")
            return False

    def _generate_sheet_name(
        self, worksheet: Worksheet, sheet_index: Optional[int]
    ) -> str:
        """Generate a meaningful sheet name based on content."""
        try:
            # Try to get name from first cell or header row
            first_cell = worksheet["A1"].value
            if first_cell and isinstance(first_cell, str) and len(first_cell) < 31:
                return self._clean_sheet_name(first_cell)

            # Try to get from first non-empty cell
            for row in worksheet.iter_rows(max_row=5, max_col=5):
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        name = self._clean_sheet_name(cell.value)
                        if name and name.lower() not in self.GENERIC_NAMES:
                            return name

            # Default to numbered name
            if sheet_index is not None:
                return f"Data Sheet {sheet_index + 1}"
            return "Data Sheet"

        except Exception:
            return "Data Sheet"

    def _clean_sheet_name(self, name: str) -> str:
        """Clean a string for use as sheet name."""
        # Remove invalid characters for Excel sheet names
        invalid_chars = [":", "\\", "/", "?", "*", "[", "]"]
        for char in invalid_chars:
            name = name.replace(char, " ")

        # Trim and limit length
        name = name.strip()[:31]
        return name

    def _make_unique_name(self, document: Any, name: str, original: str) -> str:
        """Ensure sheet name is unique."""
        if name not in document.sheetnames or name == original:
            return name

        counter = 1
        base_name = name[:27] if len(name) > 27 else name

        while f"{base_name} ({counter})" in document.sheetnames:
            counter += 1

        return f"{base_name} ({counter})"

    def _apply_table_header_fix(
        self, issue: RemediationIssue, document: Any, fix_content: str
    ) -> bool:
        """Define table headers for a data range."""
        try:
            sheet_name = issue.metadata.get("sheet_name", document.active.title)
            table_range = issue.metadata.get("table_range")

            worksheet = document[sheet_name]

            if not table_range:
                # Try to detect data range
                table_range = self._detect_data_range(worksheet)

            if not table_range:
                logger.warning("Could not determine table range")
                return False

            # Apply header formatting to first row
            start_col, start_row, end_col, end_row = self._parse_range(table_range)

            # Bold the header row
            for col in range(start_col, end_col + 1):
                cell = worksheet.cell(row=start_row, column=col)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

            # Try to create an Excel Table object
            try:
                table_name = f"Table{len(worksheet.tables) + 1}"
                table = Table(displayName=table_name, ref=table_range)
                style = TableStyleInfo(
                    name="TableStyleMedium2",
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False,
                )
                table.tableStyleInfo = style
                worksheet.add_table(table)
                logger.info(f"Created table '{table_name}' in '{sheet_name}'")
            except Exception as e:
                logger.warning(f"Could not create table object: {e}")
                # Header formatting still applied

            return True

        except Exception as e:
            logger.error(f"Error applying table header fix: {e}")
            return False

    def _detect_data_range(self, worksheet: Worksheet) -> Optional[str]:
        """Detect the data range in a worksheet."""
        try:
            if worksheet.max_row < 1 or worksheet.max_column < 1:
                return None

            # Find first row with data
            start_row = 1
            for row in range(1, min(10, worksheet.max_row + 1)):
                for col in range(1, min(10, worksheet.max_column + 1)):
                    if worksheet.cell(row=row, column=col).value:
                        start_row = row
                        break
                else:
                    continue
                break

            # Convert to Excel range format
            start_col_letter = self._get_column_letter(1)
            end_col_letter = self._get_column_letter(worksheet.max_column)

            return f"{start_col_letter}{start_row}:{end_col_letter}{worksheet.max_row}"

        except Exception:
            return None

    def _get_column_letter(self, col_num: int) -> str:
        """Convert column number to Excel letter."""
        result = ""
        while col_num > 0:
            col_num, remainder = divmod(col_num - 1, 26)
            result = chr(65 + remainder) + result
        return result

    def _parse_range(self, range_str: str) -> tuple:
        """Parse Excel range string to column/row numbers."""

        match = re.match(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", range_str)
        if not match:
            return (1, 1, 1, 1)

        def col_to_num(col):
            num = 0
            for char in col:
                num = num * 26 + (ord(char) - ord("A") + 1)
            return num

        return (
            col_to_num(match.group(1)),
            int(match.group(2)),
            col_to_num(match.group(3)),
            int(match.group(4)),
        )

    def _apply_chart_fix(
        self, issue: RemediationIssue, document: Any, alt_text: str
    ) -> bool:
        """Add alt text to a chart."""
        try:
            sheet_name = issue.metadata.get("sheet_name", document.active.title)
            chart_index = issue.metadata.get("chart_index", 0)

            worksheet = document[sheet_name]

            # Charts in openpyxl are stored in _charts
            if chart_index < len(worksheet._charts):
                chart = worksheet._charts[chart_index]

                # Set chart title if empty (as accessibility aid)
                if not chart.title and alt_text:
                    chart.title = alt_text[:50]
                    logger.info(
                        f"Applied title to chart {chart_index} in '{sheet_name}'"
                    )
                    return True

                logger.info(
                    f"Chart {chart_index} in '{sheet_name}' already has a title"
                )
                return False

            logger.warning(f"Chart not found at index {chart_index}")
            return False

        except Exception as e:
            logger.error(f"Error applying chart fix: {e}")
            return False

    def _apply_color_fix(
        self, issue: RemediationIssue, document: Any, fix_content: str
    ) -> bool:
        """Add text indicators for color-coded cells."""
        try:
            sheet_name = issue.metadata.get("sheet_name", document.active.title)
            cell_ref = issue.metadata.get("cell_ref")
            cell_row = issue.metadata.get("row")
            cell_col = issue.metadata.get("column")

            worksheet = document[sheet_name]

            if cell_ref:
                cell = worksheet[cell_ref]
            elif cell_row and cell_col:
                cell = worksheet.cell(row=cell_row, column=cell_col)
            else:
                logger.warning("No cell reference for color fix")
                return False

            # Get the color
            fill = cell.fill
            if fill and fill.fgColor and fill.fgColor.rgb:
                color = fill.fgColor.rgb
                if isinstance(color, str) and len(color) >= 6:
                    # Add text indicator based on color
                    indicator = self._get_color_indicator(color)
                    if indicator:
                        current_value = cell.value or ""
                        if indicator not in str(current_value):
                            cell.value = f"{current_value} [{indicator}]"
                            logger.info(
                                f"Added color indicator to cell {cell_ref or f'{cell_row},{cell_col}'}"
                            )
                            return True

            return False

        except Exception as e:
            logger.error(f"Error applying color fix: {e}")
            return False

    def _get_color_indicator(self, color_hex: str) -> Optional[str]:
        """Get text indicator for a color."""
        # Remove alpha if present
        if len(color_hex) == 8:
            color_hex = color_hex[2:]

        color_hex = color_hex.lower()

        # Map common colors to indicators
        color_indicators = {
            "ff0000": "Red",
            "ff6b6b": "Red",
            "00ff00": "Green",
            "00b050": "Green",
            "92d050": "Green",
            "0000ff": "Blue",
            "4472c4": "Blue",
            "ffff00": "Yellow",
            "ffc000": "Orange",
            "ff6600": "Orange",
            "7030a0": "Purple",
            "ff00ff": "Pink",
        }

        # Check for exact matches
        if color_hex in color_indicators:
            return color_indicators[color_hex]

        # Try to determine by RGB components
        try:
            r = int(color_hex[0:2], 16)
            g = int(color_hex[2:4], 16)
            b = int(color_hex[4:6], 16)

            if r > 200 and g < 100 and b < 100:
                return "Red"
            elif g > 200 and r < 100 and b < 100:
                return "Green"
            elif b > 200 and r < 100 and g < 100:
                return "Blue"
            elif r > 200 and g > 200 and b < 100:
                return "Yellow"
            elif r > 200 and g > 100 and b < 100:
                return "Orange"

        except Exception:
            pass

        return None

    def _apply_navigation_fix(
        self, issue: RemediationIssue, document: Any, fix_content: str
    ) -> bool:
        """Add navigation aids (frozen panes) to worksheet."""
        try:
            sheet_name = issue.metadata.get("sheet_name", document.active.title)
            worksheet = document[sheet_name]

            # Freeze first row (typically headers)
            worksheet.freeze_panes = "A2"

            logger.info(f"Froze header row in '{sheet_name}'")
            return True

        except Exception as e:
            logger.error(f"Error applying navigation fix: {e}")
            return False

    def _get_rule_based_fix(
        self, issue: RemediationIssue, document: Any
    ) -> Optional[str]:
        """Get a rule-based fix for an issue."""
        if issue.category == IssueCategory.ALT_TEXT:
            # Use pre-generated alt text from the scanner if available
            # Scanner stores as "suggested_alt_text", check both keys
            generated_alt = issue.metadata.get(
                "suggested_alt_text"
            ) or issue.metadata.get("generated_alt_text")
            if generated_alt:
                return generated_alt
            # Use fix_suggestion if available
            if issue.fix_suggestion:
                return issue.fix_suggestion
            # Return None to let AI generation handle it in _generate_fix()
            return None

        if issue.category == IssueCategory.SHEET:
            sheet_name = issue.metadata.get("sheet_name", "")
            sheet_index = issue.metadata.get("sheet_index")
            if document and sheet_name in document.sheetnames:
                return self._generate_sheet_name(document[sheet_name], sheet_index)
            return f"Data Sheet {(sheet_index or 0) + 1}"

        if issue.category == IssueCategory.TABLE:
            return "header_row"

        if issue.category == IssueCategory.NAVIGATION:
            return "freeze_panes"

        return None

    def _get_ai_generated_fix(
        self, issue: RemediationIssue, document: Any
    ) -> Optional[str]:
        """Get an AI-generated fix for an issue."""
        if not self.ai_client:
            return None

        try:
            self.result.ai_calls_made += 1

            if issue.category == IssueCategory.SHEET:
                return self._generate_sheet_name_with_ai(issue, document)

            if issue.category == IssueCategory.CHART:
                return self._generate_chart_description_with_ai(issue, document)

            return None

        except Exception as e:
            logger.error(f"AI generation failed: {e}")
            return None

    def _generate_sheet_name_with_ai(
        self, issue: RemediationIssue, document: Any
    ) -> Optional[str]:
        """Generate a sheet name using AI based on content."""
        sheet_name = issue.metadata.get("sheet_name", document.active.title)

        # Get sample data from sheet
        content_sample = ""
        try:
            worksheet = document[sheet_name]
            cells = []
            for row in worksheet.iter_rows(max_row=5, max_col=5):
                for cell in row:
                    if cell.value:
                        cells.append(str(cell.value)[:50])
            content_sample = ", ".join(cells[:20])
        except Exception:
            pass

        from ...utils.security import sanitize_for_prompt

        safe_sheet_name = sanitize_for_prompt(sheet_name, max_length=100)
        safe_content = (
            sanitize_for_prompt(content_sample, max_length=300)
            if content_sample
            else "No content available"
        )

        prompt = f"""Generate a concise, descriptive name for an Excel worksheet.

Current name: "{safe_sheet_name}"
Sample content: {safe_content}

Requirements:
- Keep it short (max 31 characters)
- Make it descriptive of the data
- Use title case
- Don't include special characters: : \\ / ? * [ ]

Generate only the sheet name, nothing else:"""

        try:
            if hasattr(self.ai_client, "generate_text_sync"):
                result = self.ai_client.generate_text_sync(
                    prompt=prompt,
                    max_tokens=100,
                    temperature=0.3,
                )
                if result.get("success") and result.get("content"):
                    return self._clean_sheet_name(
                        result["content"].strip().strip("\"'")
                    )
        except Exception as e:
            logger.error(f"AI sheet name generation failed: {e}")

        return None

    def _generate_chart_description_with_ai(
        self, issue: RemediationIssue, document: Any
    ) -> Optional[str]:
        """Generate chart description using AI."""
        sheet_name = issue.metadata.get("sheet_name", document.active.title)

        # Get data context
        content_sample = ""
        try:
            worksheet = document[sheet_name]
            cells = []
            for row in worksheet.iter_rows(max_row=10, max_col=10):
                for cell in row:
                    if cell.value:
                        cells.append(str(cell.value)[:30])
            content_sample = ", ".join(cells[:30])
        except Exception:
            pass

        from ...utils.security import sanitize_for_prompt

        safe_sheet_name = sanitize_for_prompt(sheet_name, max_length=100)
        safe_content = (
            sanitize_for_prompt(content_sample, max_length=300)
            if content_sample
            else "No data available"
        )

        prompt = f"""Generate a concise description for a chart in an Excel spreadsheet.

Sheet: "{safe_sheet_name}"
Nearby data: {safe_content}

Requirements:
- Be concise (under 100 characters)
- Describe what the chart likely shows
- Focus on the data relationship

Generate only the description, nothing else:"""

        try:
            if hasattr(self.ai_client, "generate_text_sync"):
                result = self.ai_client.generate_text_sync(
                    prompt=prompt,
                    max_tokens=150,
                    temperature=0.3,
                )
                if result.get("success") and result.get("content"):
                    return result["content"].strip().strip("\"'")[:100]
        except Exception as e:
            logger.error(f"AI chart description generation failed: {e}")

        # Fail closed: never emit a placeholder description counted as a fix.
        # Returning None routes this issue to the human review queue (WCAG 1.1.1).
        return None

    def _calculate_scores(self):
        """Calculate compliance scores for the remediation."""
        if self.result.total_issues > 0:
            severity_penalties = {
                IssueSeverity.CRITICAL: 15,
                IssueSeverity.HIGH: 10,
                IssueSeverity.MEDIUM: 5,
                IssueSeverity.LOW: 2,
            }

            total_penalty = sum(
                severity_penalties.get(issue.severity, 5) for issue in self.issues
            )
            self.result.original_compliance_score = max(0, 100 - total_penalty)

            fixed_penalty_reduction = sum(
                severity_penalties.get(fixed.severity, 5)
                for fixed in self.result.fixed_issues
            )
            remaining_penalty = total_penalty - fixed_penalty_reduction
            self.result.remediated_compliance_score = max(0, 100 - remaining_penalty)

            self.result.improvement = (
                self.result.remediated_compliance_score
                - self.result.original_compliance_score
            )
