"""
Excel Spreadsheet (.xlsx) Accessibility Scanner Module

This module provides functionality to:
1. Parse XLSX files and analyze structure
2. Check sheet names (meaningful vs "Sheet1")
3. Detect table headers (defined vs missing)
4. Check alt text on charts/images
5. Analyze cell merges for accessibility
6. Detect color-only information
7. Check for frozen panes (navigation aid)
8. Verify named ranges for screen readers
9. Generate AI-powered chart descriptions
10. Batch process entire directories

Spreadsheets need accessibility too - this module ensures data is accessible to all.
"""

from typing import List, Dict, Optional, Tuple, Any
from pydantic import BaseModel, computed_field
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
import os
import tempfile
from pathlib import Path

from src.utils.async_helpers import run_async_from_sync
from PIL import Image
from io import BytesIO
import logging
import re

from src.education.color_blindness_simulator import (
    ColorBlindnessSimulator,
    ColorBlindnessAnalysisResult,
)

logger = logging.getLogger(__name__)


class SheetNameIssue(BaseModel):
    """Sheet name accessibility issue"""

    sheet_name: str
    sheet_index: int
    issue_type: str  # generic_name, too_long, special_chars
    suggested_fix: str


class TableHeaderIssue(BaseModel):
    """Table header accessibility issue"""

    sheet_name: str
    table_range: str
    issue_type: str  # missing_header, ambiguous_header
    row_count: int
    column_count: int
    suggested_fix: str


class ChartIssue(BaseModel):
    """Chart accessibility issue"""

    sheet_name: str
    chart_index: int
    chart_type: str
    has_alt_text: bool
    existing_alt_text: Optional[str] = None
    suggested_alt_text: Optional[str] = None
    detailed_description: Optional[str] = None
    data_summary: Optional[str] = None


class ImageIssue(BaseModel):
    """Embedded image accessibility issue"""

    sheet_name: str
    image_index: int
    has_alt_text: bool
    existing_alt_text: Optional[str] = None
    suggested_alt_text: Optional[str] = None
    detected_image_type: Optional[str] = None
    is_decorative: bool = False
    # Alt text validation (for images WITH alt text)
    alt_text_validated: bool = False  # Whether AI validation was performed
    alt_text_accurate: Optional[bool] = None  # Whether existing alt text is accurate
    alt_text_issues: Optional[List[str]] = None  # Specific problems found
    validation_score: Optional[float] = None  # Accuracy score 0-1


class MergeCellIssue(BaseModel):
    """Merged cell accessibility issue"""

    sheet_name: str
    merge_range: str
    rows_merged: int
    cols_merged: int
    issue_type: str  # large_merge, header_merge, data_merge
    suggested_fix: str


class ColorOnlyIssue(BaseModel):
    """Color-only information issue"""

    sheet_name: str
    cell_range: str
    issue_type: str  # color_coding, no_legend
    colors_used: List[str]
    suggested_fix: str


class NavigationIssue(BaseModel):
    """Navigation accessibility issue"""

    sheet_name: str
    issue_type: str  # no_freeze, no_filter, large_dataset
    suggested_fix: str


class ConditionalFormatIssue(BaseModel):
    """Conditional formatting accessibility issue (WCAG 1.4.1)"""

    sheet_name: str
    cell_range: str
    rule_type: str  # colorScale, dataBar, iconSet, cellIs
    uses_color_only: bool
    has_text_alternative: bool = False
    color_count: int = 0
    recommendation: str


class PivotTableIssue(BaseModel):
    """Pivot table accessibility issue (WCAG 1.3.1)"""

    sheet_name: str
    pivot_name: str
    pivot_location: str  # Cell range of pivot table
    row_fields: List[str] = []
    column_fields: List[str] = []
    value_fields: List[str] = []
    has_field_labels: bool = True
    has_grand_totals: bool = False
    issue_type: str  # complex_structure, missing_labels, nested_headers
    recommendations: List[str]


class ContrastIssue(BaseModel):
    """Text contrast accessibility issue (WCAG 1.4.3)"""

    sheet_name: str
    cell_reference: str  # e.g., "A1", "B5"
    text_preview: str  # First 30 chars of cell content
    foreground_color: str  # Hex color
    background_color: str  # Hex color
    contrast_ratio: float
    wcag_aa_pass: bool  # 4.5:1 for normal text
    suggested_fix: str


class SheetAnalysis(BaseModel):
    """Analysis results for a single sheet"""

    sheet_name: str
    row_count: int
    column_count: int
    has_data: bool
    has_tables: bool
    has_charts: bool
    has_images: bool
    has_merged_cells: bool
    has_frozen_panes: bool
    table_header_issues: List[TableHeaderIssue]
    chart_issues: List[ChartIssue]
    image_issues: List[ImageIssue]
    merge_issues: List[MergeCellIssue]
    color_issues: List[ColorOnlyIssue]
    navigation_issues: List[NavigationIssue]
    contrast_issues: List[ContrastIssue] = []  # Text contrast issues (WCAG 1.4.3)
    conditional_format_issues: List[ConditionalFormatIssue] = []  # CF accessibility
    pivot_table_issues: List[PivotTableIssue] = []  # Pivot table accessibility


class XlsxProcessingResult(BaseModel):
    """Result of Excel processing operation"""

    file_path: str
    file_name: str
    total_sheets: int
    total_rows: int
    total_charts: int
    total_images: int
    sheet_name_issues: List[SheetNameIssue]
    sheets: List[SheetAnalysis]
    summary: Dict[str, int]
    compliance_score: float
    remediation_suggestions: List[str]
    # Color vision deficiency analysis
    cvd_analysis: Optional[List[ColorBlindnessAnalysisResult]] = None

    @computed_field
    @property
    def issues(self) -> List[Dict[str, Any]]:
        """Combined list of all issues for API compatibility.

        The demo routes expect a single 'issues' array, but XLSX processor
        stores issues in separate category arrays across sheets. This computed
        field combines them all into a unified format.
        """
        all_issues: List[Dict[str, Any]] = []

        # Map severity from issue types
        def get_severity(issue_type: str) -> str:
            critical_types = ["missing_alt_text", "missing_header"]
            high_types = ["generic_name", "color_only", "navigation"]
            medium_types = ["merge_cell", "contrast"]
            if any(t in issue_type for t in critical_types):
                return "critical"
            elif any(t in issue_type for t in high_types):
                return "high"
            elif any(t in issue_type for t in medium_types):
                return "medium"
            return "low"

        # Sheet name issues (document-level)
        for issue in self.sheet_name_issues:
            if issue.issue_type == "generic_name":
                desc = f"Sheet name '{issue.sheet_name}' is generic and not descriptive"
            elif issue.issue_type == "too_long":
                desc = f"Sheet name '{issue.sheet_name}' exceeds recommended length"
            else:
                desc = f"Sheet name issue: {issue.issue_type}"
            all_issues.append(
                {
                    "id": f"sheet_name_{len(all_issues)}",
                    "category": "sheet_name",
                    "severity": get_severity(issue.issue_type),
                    "title": f"Generic Sheet Name: {issue.sheet_name}",
                    "description": desc,
                    "location": f"Sheet '{issue.sheet_name}'",
                    "wcag_criterion": "WCAG 2.4.2",
                    "suggested_fix": issue.suggested_fix,
                }
            )

        # Per-sheet issues
        for sheet in self.sheets:
            sheet_loc = f"Sheet '{sheet.sheet_name}'"

            # Table header issues
            for issue in sheet.table_header_issues:
                desc = f"Table at {issue.table_range} ({issue.row_count}×{issue.column_count}) is missing header row designation"
                all_issues.append(
                    {
                        "id": f"table_{len(all_issues)}",
                        "category": "table",
                        "severity": get_severity(issue.issue_type),
                        "title": issue.issue_type.replace("_", " ").title(),
                        "description": desc,
                        "location": f"{sheet_loc}, {issue.table_range}",
                        "wcag_criterion": "WCAG 1.3.1",
                        "suggested_fix": issue.suggested_fix,
                    }
                )

            # Chart issues
            for issue in sheet.chart_issues:
                if issue.has_alt_text:
                    desc = f"{issue.chart_type} chart has alt text that may need review"
                else:
                    desc = f"{issue.chart_type} chart is missing alternative text description"
                all_issues.append(
                    {
                        "id": f"chart_{len(all_issues)}",
                        "category": "alt_text",
                        "severity": "critical" if not issue.has_alt_text else "medium",
                        "title": (
                            "Chart Missing Description"
                            if not issue.has_alt_text
                            else "Chart Description Review"
                        ),
                        "description": desc,
                        "location": f"{sheet_loc}, Chart {issue.chart_index + 1}",
                        "wcag_criterion": "WCAG 1.1.1",
                        "suggested_fix": issue.suggested_alt_text
                        or "Add descriptive alt text for the chart",
                        "ai_generated": issue.suggested_alt_text is not None,
                        "generated_alt_text": issue.suggested_alt_text,
                    }
                )

            # Image issues
            for issue in sheet.image_issues:
                if issue.has_alt_text:
                    # Image has alt text - check if validation was performed
                    if issue.alt_text_validated and not issue.alt_text_accurate:
                        # Alt text was validated and found to be inaccurate
                        issues_list = (
                            ", ".join(issue.alt_text_issues)
                            if issue.alt_text_issues
                            else "does not accurately describe the image"
                        )
                        existing_preview = (
                            issue.existing_alt_text[:50] + "..."
                            if issue.existing_alt_text
                            and len(issue.existing_alt_text) > 50
                            else issue.existing_alt_text
                        )
                        desc = f"Alt text '{existing_preview}' is inaccurate: {issues_list}"
                        title = "Inaccurate Alt Text"
                        severity = (
                            "high"
                            if issue.validation_score and issue.validation_score < 0.5
                            else "medium"
                        )
                    else:
                        existing_preview = (
                            issue.existing_alt_text[:50] + "..."
                            if issue.existing_alt_text
                            and len(issue.existing_alt_text) > 50
                            else issue.existing_alt_text
                        )
                        desc = f"Image has alt text that may need review: '{existing_preview}'"
                        title = "Alt Text Review"
                        severity = "low"
                else:
                    if issue.is_decorative:
                        desc = f'Image {issue.image_index + 1} is decorative and should have empty alt text (alt="")'
                        title = "Decorative Image — Needs Empty Alt Text"
                        severity = "low"
                    else:
                        desc = (
                            f"Image {issue.image_index + 1} is missing alternative text"
                        )
                        title = "Missing Alt Text"
                        severity = "critical"
                all_issues.append(
                    {
                        "id": f"image_{len(all_issues)}",
                        "category": "alt_text",
                        "severity": severity,
                        "title": title,
                        "description": desc,
                        "location": f"{sheet_loc}, Image {issue.image_index + 1}",
                        "wcag_criterion": "WCAG 1.1.1",
                        "suggested_fix": issue.suggested_alt_text
                        or (
                            "Mark as decorative (set empty alt text)"
                            if issue.is_decorative
                            else "Add descriptive alt text"
                        ),
                        "ai_generated": issue.suggested_alt_text is not None,
                        "generated_alt_text": issue.suggested_alt_text or None,
                        "is_decorative": issue.is_decorative,
                        # Validation metadata
                        "alt_text_validated": issue.alt_text_validated,
                        "alt_text_accurate": issue.alt_text_accurate,
                        "alt_text_issues": issue.alt_text_issues,
                        "validation_score": issue.validation_score,
                    }
                )

            # Merge cell issues
            for issue in sheet.merge_issues:
                desc = f"Cells {issue.merge_range} are merged ({issue.rows_merged}×{issue.cols_merged}), which may confuse screen readers"
                all_issues.append(
                    {
                        "id": f"merge_{len(all_issues)}",
                        "category": "structure",
                        "severity": get_severity(issue.issue_type),
                        "title": issue.issue_type.replace("_", " ").title(),
                        "description": desc,
                        "location": f"{sheet_loc}, {issue.merge_range}",
                        "wcag_criterion": "WCAG 1.3.1",
                        "suggested_fix": issue.suggested_fix,
                    }
                )

            # Color-only issues
            for issue in sheet.color_issues:
                colors_str = ", ".join(issue.colors_used[:3])
                if len(issue.colors_used) > 3:
                    colors_str += f" (+{len(issue.colors_used) - 3} more)"
                desc = f"Cells at {issue.cell_range} use color alone to convey information (colors: {colors_str})"
                all_issues.append(
                    {
                        "id": f"color_{len(all_issues)}",
                        "category": "color",
                        "severity": get_severity(issue.issue_type),
                        "title": issue.issue_type.replace("_", " ").title(),
                        "description": desc,
                        "location": f"{sheet_loc}, {issue.cell_range}",
                        "wcag_criterion": "WCAG 1.4.1",
                        "suggested_fix": issue.suggested_fix,
                    }
                )

            # Navigation issues
            for issue in sheet.navigation_issues:
                if issue.issue_type == "no_freeze":
                    desc = "Sheet does not have frozen panes, making navigation difficult for large datasets"
                elif issue.issue_type == "large_dataset":
                    desc = "Sheet contains a large dataset without navigation aids"
                else:
                    desc = f"Navigation issue: {issue.issue_type}"
                all_issues.append(
                    {
                        "id": f"nav_{len(all_issues)}",
                        "category": "navigation",
                        "severity": get_severity(issue.issue_type),
                        "title": issue.issue_type.replace("_", " ").title(),
                        "description": desc,
                        "location": sheet_loc,
                        "wcag_criterion": "WCAG 2.4.3",
                        "suggested_fix": issue.suggested_fix,
                    }
                )

            # Contrast issues
            for issue in sheet.contrast_issues:
                desc = f"Text '{issue.text_preview}' has insufficient contrast ({issue.contrast_ratio:.1f}:1, needs 4.5:1)"
                all_issues.append(
                    {
                        "id": f"contrast_{len(all_issues)}",
                        "category": "contrast",
                        "severity": "high" if issue.contrast_ratio < 3.0 else "medium",
                        "title": f"Low Contrast ({issue.contrast_ratio:.1f}:1)",
                        "description": desc,
                        "location": f"{sheet_loc}, {issue.cell_reference}",
                        "wcag_criterion": "WCAG 1.4.3",
                        "suggested_fix": issue.suggested_fix,
                    }
                )

            # Conditional formatting issues
            for issue in sheet.conditional_format_issues:
                if issue.uses_color_only:
                    desc = f"Conditional formatting at {issue.cell_range} uses {issue.rule_type} with color-only indicators"
                else:
                    desc = f"Conditional formatting at {issue.cell_range} uses {issue.rule_type}"
                all_issues.append(
                    {
                        "id": f"cf_{len(all_issues)}",
                        "category": "color",
                        "severity": "medium" if issue.uses_color_only else "low",
                        "title": f"Conditional Format: {issue.rule_type}",
                        "description": desc,
                        "location": f"{sheet_loc}, {issue.cell_range}",
                        "wcag_criterion": "WCAG 1.4.1",
                        "suggested_fix": issue.recommendation,
                    }
                )

            # Pivot table issues
            for issue in sheet.pivot_table_issues:
                field_info = []
                if issue.row_fields:
                    field_info.append(f"rows: {', '.join(issue.row_fields[:3])}")
                if issue.column_fields:
                    field_info.append(f"columns: {', '.join(issue.column_fields[:3])}")
                fields_str = (
                    "; ".join(field_info) if field_info else "complex structure"
                )

                desc = f"Pivot table '{issue.pivot_name}' at {issue.pivot_location} has accessibility concerns ({fields_str})"
                all_issues.append(
                    {
                        "id": f"pivot_{len(all_issues)}",
                        "category": "structure",
                        "severity": (
                            "medium"
                            if issue.issue_type == "complex_structure"
                            else "low"
                        ),
                        "title": f"Pivot Table: {issue.issue_type.replace('_', ' ').title()}",
                        "description": desc,
                        "location": f"{sheet_loc}, {issue.pivot_location}",
                        "wcag_criterion": "WCAG 1.3.1",
                        "suggested_fix": "; ".join(issue.recommendations[:2]),
                    }
                )

        return all_issues


class XlsxProcessor:
    """Process Excel files for accessibility compliance"""

    # Generic sheet names to flag
    GENERIC_NAMES = [
        "sheet1",
        "sheet2",
        "sheet3",
        "sheet",
        "data",
        "new sheet",
        "worksheet",
    ]

    def __init__(
        self,
        generate_chart_descriptions: bool = False,
        generate_alt_text: bool = False,
        validate_alt_text: bool = False,
        simulate_color_blindness: bool = False,
        progress_callback: callable = None,
    ):
        self.generate_chart_descriptions = generate_chart_descriptions
        self.generate_alt_text = generate_alt_text
        self.validate_alt_text = validate_alt_text
        self.progress_callback = progress_callback
        self.image_generator = None
        # Color vision deficiency simulation
        self.simulate_color_blindness = simulate_color_blindness
        self.cvd_simulator = (
            ColorBlindnessSimulator() if simulate_color_blindness else None
        )

        # Lazy import image generator if needed
        if (
            self.generate_chart_descriptions
            or self.generate_alt_text
            or self.validate_alt_text
        ):
            try:
                from .image_alt_text import ImageAltTextGenerator

                self.image_generator = ImageAltTextGenerator(
                    allow_legacy_transport=True
                )
            except Exception as e:
                logger.warning(
                    f"[XlsxProcessor] Could not initialize ImageAltTextGenerator: {e}"
                )
                self.generate_chart_descriptions = False
                self.generate_alt_text = False
                self.validate_alt_text = False

    def process_xlsx(
        self, file_path: str, original_filename: str = None
    ) -> XlsxProcessingResult:
        """
        Process an Excel file and check accessibility

        Args:
            file_path: Path to XLSX file
            original_filename: Optional original filename

        Returns:
            XlsxProcessingResult with all accessibility issues
        """
        wb = load_workbook(file_path, data_only=True)
        file_name = original_filename or os.path.basename(file_path)
        total_sheets = len(wb.sheetnames)

        # Report initial progress
        if self.progress_callback:
            self.progress_callback(0, total_sheets + 2, "Loading workbook...")

        # Extract workbook context
        workbook_context = self._extract_workbook_context(wb, file_name)

        if self.progress_callback:
            self.progress_callback(1, total_sheets + 2, "Checking sheet names...")

        # Check sheet names
        sheet_name_issues = self._check_sheet_names(wb)

        # Analyze each sheet
        sheets_analysis = []
        total_rows = 0
        total_charts = 0
        total_images = 0

        for sheet_idx, sheet_name in enumerate(wb.sheetnames):
            # Report progress per sheet
            if self.progress_callback:
                self.progress_callback(
                    sheet_idx + 2,
                    total_sheets + 2,
                    f"Analyzing sheet '{sheet_name}' ({sheet_idx + 1} of {total_sheets})...",
                )
            ws = wb[sheet_name]
            sheet_context = self._extract_sheet_context(
                ws, sheet_name, workbook_context
            )

            analysis = self._analyze_sheet(ws, sheet_name, sheet_context)
            sheets_analysis.append(analysis)

            total_rows += analysis.row_count
            total_charts += len([c for c in analysis.chart_issues])
            total_images += len([i for i in analysis.image_issues])

        # Calculate summary
        summary = self._calculate_summary(sheet_name_issues, sheets_analysis)

        # Calculate compliance score
        total_elements = total_rows + total_charts + total_images + len(wb.sheetnames)
        compliance_score = self._calculate_compliance_score(summary, total_elements)

        # Generate remediation suggestions
        remediation_suggestions = self._generate_remediation_suggestions(summary)

        # Analyze color vision deficiency accessibility if enabled
        cvd_analysis = None
        if self.simulate_color_blindness and self.cvd_simulator:
            logger.info("[XlsxProcessor] Running CVD accessibility analysis...")
            cvd_analysis = self._analyze_cvd_accessibility(wb)
            if cvd_analysis:
                logger.info(
                    f"[XlsxProcessor] CVD analysis complete: {len(cvd_analysis)} color pairs tested"
                )

        wb.close()

        return XlsxProcessingResult(
            file_path=file_path,
            file_name=file_name,
            total_sheets=len(wb.sheetnames),
            total_rows=total_rows,
            total_charts=total_charts,
            total_images=total_images,
            sheet_name_issues=sheet_name_issues,
            sheets=sheets_analysis,
            summary=summary,
            compliance_score=compliance_score,
            remediation_suggestions=remediation_suggestions,
            cvd_analysis=cvd_analysis,
        )

    def _extract_workbook_context(self, wb, filename: str) -> Dict:
        """Extract workbook-level context for AI"""
        context = {
            "filename": filename,
            "sheet_names": wb.sheetnames,
            "total_sheets": len(wb.sheetnames),
            "has_multiple_sheets": len(wb.sheetnames) > 1,
        }

        # Try to infer workbook purpose from sheet names and data
        purpose_keywords = []
        for name in wb.sheetnames:
            purpose_keywords.extend(name.lower().split())

        context["purpose_keywords"] = list(set(purpose_keywords))

        return context

    def _extract_sheet_context(
        self, ws: Worksheet, sheet_name: str, workbook_context: Dict
    ) -> Dict:
        """Extract sheet-level context for AI"""
        context = {
            "sheet_name": sheet_name,
            "workbook_context": workbook_context,
            "headers": [],
            "data_sample": [],
        }

        # Get header row (usually row 1)
        if ws.max_row > 0:
            headers = []
            for col in range(1, min(ws.max_column + 1, 20)):  # Limit to 20 columns
                cell = ws.cell(row=1, column=col)
                if cell.value:
                    headers.append(str(cell.value)[:50])
            context["headers"] = headers

        # Get sample data from first few rows
        if ws.max_row > 1:
            for row_num in range(2, min(ws.max_row + 1, 5)):  # Rows 2-4
                row_data = []
                for col in range(1, min(ws.max_column + 1, 10)):
                    cell = ws.cell(row=row_num, column=col)
                    if cell.value:
                        row_data.append(str(cell.value)[:30])
                if row_data:
                    context["data_sample"].append(row_data)

        return context

    def _check_sheet_names(self, wb) -> List[SheetNameIssue]:
        """Check sheet names for accessibility"""
        issues = []

        for idx, name in enumerate(wb.sheetnames):
            # Check for generic names
            if name.lower().strip() in self.GENERIC_NAMES:
                issues.append(
                    SheetNameIssue(
                        sheet_name=name,
                        sheet_index=idx,
                        issue_type="generic_name",
                        suggested_fix=f'Rename "{name}" to describe its content (e.g., "Q1 Sales Data", "Employee Directory")',
                    )
                )

            # Check for very long names
            elif len(name) > 31:  # Excel limit is 31 chars
                issues.append(
                    SheetNameIssue(
                        sheet_name=name,
                        sheet_index=idx,
                        issue_type="too_long",
                        suggested_fix=f'Sheet name "{name[:20]}..." is too long. Keep under 31 characters.',
                    )
                )

            # Check for special characters that may cause issues
            elif re.search(r"[\\/*?:\[\]]", name):
                issues.append(
                    SheetNameIssue(
                        sheet_name=name,
                        sheet_index=idx,
                        issue_type="special_chars",
                        suggested_fix=f'Remove special characters from "{name}". Use only letters, numbers, and spaces.',
                    )
                )

        return issues

    def _analyze_sheet(
        self, ws: Worksheet, sheet_name: str, context: Dict
    ) -> SheetAnalysis:
        """Analyze a single sheet for accessibility issues"""

        # Get dimensions
        row_count = ws.max_row or 0
        col_count = ws.max_column or 0
        has_data = row_count > 0 and col_count > 0

        # Check for tables
        has_tables = len(ws.tables) > 0

        # Check for charts
        charts = ws._charts if hasattr(ws, "_charts") else []
        has_charts = len(charts) > 0

        # Check for images
        images = ws._images if hasattr(ws, "_images") else []
        has_images = len(images) > 0

        # Check merged cells
        merged_ranges = list(ws.merged_cells.ranges) if ws.merged_cells else []
        has_merged_cells = len(merged_ranges) > 0

        # Check frozen panes
        has_frozen_panes = ws.freeze_panes is not None

        # Analyze issues
        table_header_issues = self._check_table_headers(ws, sheet_name, context)
        chart_issues = self._check_charts(ws, sheet_name, charts, context)
        image_issues = self._check_images(ws, sheet_name, images, context)
        merge_issues = self._check_merged_cells(ws, sheet_name, merged_ranges)
        color_issues = self._check_color_coding(ws, sheet_name)
        navigation_issues = self._check_navigation(
            ws, sheet_name, row_count, col_count, has_frozen_panes
        )
        contrast_issues = self._check_text_contrast(ws, sheet_name)
        conditional_format_issues = self._analyze_conditional_formatting(ws, sheet_name)
        pivot_table_issues = self._analyze_pivot_tables(ws, sheet_name)

        return SheetAnalysis(
            sheet_name=sheet_name,
            row_count=row_count,
            column_count=col_count,
            has_data=has_data,
            has_tables=has_tables,
            has_charts=has_charts,
            has_images=has_images,
            has_merged_cells=has_merged_cells,
            has_frozen_panes=has_frozen_panes,
            table_header_issues=table_header_issues,
            chart_issues=chart_issues,
            image_issues=image_issues,
            merge_issues=merge_issues,
            color_issues=color_issues,
            navigation_issues=navigation_issues,
            contrast_issues=contrast_issues,
            conditional_format_issues=conditional_format_issues,
            pivot_table_issues=pivot_table_issues,
        )

    def _check_table_headers(
        self, ws: Worksheet, sheet_name: str, context: Dict
    ) -> List[TableHeaderIssue]:
        """Check for proper table headers"""
        issues = []

        # If sheet has data but no defined tables
        if ws.max_row > 1 and ws.max_column > 0 and len(ws.tables) == 0:
            # Check if first row looks like headers
            first_row_values = []
            has_header_formatting = False

            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col)
                if cell.value:
                    first_row_values.append(str(cell.value))

                # Check for header-like formatting (bold, different background)
                if cell.font and cell.font.bold:
                    has_header_formatting = True
                if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb:
                    if cell.fill.start_color.rgb != "00000000":  # Not transparent
                        has_header_formatting = True

            # If data looks tabular but no explicit table/header defined
            if len(first_row_values) > 1 and not has_header_formatting:
                issues.append(
                    TableHeaderIssue(
                        sheet_name=sheet_name,
                        table_range=f"A1:{get_column_letter(ws.max_column)}{ws.max_row}",
                        issue_type="missing_header",
                        row_count=ws.max_row,
                        column_count=ws.max_column,
                        suggested_fix="Define first row as header. Format cells as bold or use Insert > Table to create a proper table with headers.",
                    )
                )

        # Check existing tables for header issues
        for table in ws.tables.values():
            if not table.headerRowCount or table.headerRowCount == 0:
                issues.append(
                    TableHeaderIssue(
                        sheet_name=sheet_name,
                        table_range=table.ref,
                        issue_type="missing_header",
                        row_count=0,
                        column_count=0,
                        suggested_fix=f'Table {table.name} has no header row defined. Check "My table has headers" in Table Design.',
                    )
                )

        return issues

    def _check_charts(
        self, ws: Worksheet, sheet_name: str, charts: List, context: Dict
    ) -> List[ChartIssue]:
        """Check charts for alt text and descriptions"""
        issues = []

        for idx, chart in enumerate(charts):
            chart_type = type(chart).__name__.replace("Chart", "")

            # Check for alt text/title
            has_alt = False

            if hasattr(chart, "title") and chart.title:
                has_alt = True

            suggested_alt = None
            detailed_desc = None
            data_summary = None

            # Generate chart description with AI if enabled
            if (
                not has_alt
                and self.generate_chart_descriptions
                and self.image_generator
            ):
                try:
                    # Build context for chart
                    chart_context = self._build_chart_context(chart, context)

                    # For charts, we generate a text-based description
                    # since we can't easily export chart as image from openpyxl
                    suggested_alt, detailed_desc, data_summary = (
                        self._generate_chart_description(
                            chart, chart_type, chart_context
                        )
                    )
                except Exception as e:
                    logger.warning(
                        f"[XlsxProcessor] Chart description generation failed: {e}"
                    )

            if not has_alt:
                issues.append(
                    ChartIssue(
                        sheet_name=sheet_name,
                        chart_index=idx,
                        chart_type=chart_type,
                        has_alt_text=False,
                        existing_alt_text=None,
                        suggested_alt_text=suggested_alt,
                        detailed_description=detailed_desc,
                        data_summary=data_summary,
                    )
                )

        return issues

    def _build_chart_context(self, chart, context: Dict) -> str:
        """Build context string for chart description"""
        parts = []

        parts.append(f"Sheet: {context.get('sheet_name', 'Unknown')}")

        if context.get("headers"):
            parts.append(f"Data headers: {', '.join(context['headers'][:5])}")

        if context.get("data_sample"):
            parts.append("Sample data:")
            for row in context["data_sample"][:2]:
                parts.append(f"  {', '.join(row[:5])}")

        return "\n".join(parts)

    def _generate_chart_description(
        self, chart, chart_type: str, context: str
    ) -> Tuple[str, str, str]:
        """Generate description for a chart"""
        # Basic description based on chart type
        type_descriptions = {
            "Bar": "bar chart comparing values across categories",
            "Line": "line chart showing trends over time",
            "Pie": "pie chart showing proportional distribution",
            "Area": "area chart displaying cumulative values over time",
            "Scatter": "scatter plot showing correlation between variables",
            "Doughnut": "doughnut chart showing proportional distribution",
        }

        base_desc = type_descriptions.get(chart_type, f"{chart_type} chart")

        # Try to get data range info
        data_summary = "Data visualization"
        if hasattr(chart, "series") and chart.series:
            series_count = len(chart.series)
            data_summary = f"Chart with {series_count} data series"

        suggested_alt = f"A {base_desc}"
        detailed_desc = f"{suggested_alt}. {data_summary}. Context: {context[:200]}"

        return suggested_alt, detailed_desc, data_summary

    def _check_images(
        self, ws: Worksheet, sheet_name: str, images: List, context: Dict
    ) -> List[ImageIssue]:
        """Check embedded images for alt text"""
        issues = []

        for idx, img in enumerate(images):
            # openpyxl Image objects don't have great alt text support
            # We'll flag all images as needing review
            has_alt = False
            existing_alt_text = None

            # Try to get description if available
            if hasattr(img, "desc") and img.desc:
                has_alt = True
                existing_alt_text = img.desc

            suggested_alt = None
            detected_type = None
            is_decorative = False

            # Generate alt text with AI if enabled (for images WITHOUT alt text)
            if not has_alt and self.generate_alt_text and self.image_generator:
                try:
                    image_path = self._extract_image(img)
                    if image_path:
                        img_context = f"Image in Excel sheet '{sheet_name}'"
                        if context.get("headers"):
                            img_context += (
                                f". Sheet headers: {', '.join(context['headers'][:5])}"
                            )

                        # Detect image type
                        type_result = run_async_from_sync(
                            self.image_generator.detect_image_type(
                                image_path=image_path, context=img_context
                            )
                        )

                        if type_result.get("success"):
                            detected_type = type_result.get("image_type", "informative")
                            is_decorative = type_result.get("is_decorative", False)

                            if not is_decorative:
                                result = run_async_from_sync(
                                    self.image_generator.generate_alt_text(
                                        image_path=image_path,
                                        context=img_context,
                                        educational_context=True,
                                    )
                                )
                                if result.get("success"):
                                    suggested_alt = result.get("alt_text")

                        try:
                            os.unlink(image_path)
                        except OSError as e:
                            logger.debug(
                                f"[XlsxProcessor] Could not remove temp image "
                                f"'{image_path}': {e}"
                            )
                except Exception as e:
                    logger.warning(
                        f"[XlsxProcessor] Image alt text generation failed: {e}"
                    )

            if not has_alt:
                issues.append(
                    ImageIssue(
                        sheet_name=sheet_name,
                        image_index=idx,
                        has_alt_text=False,
                        existing_alt_text=None,
                        suggested_alt_text=suggested_alt,
                        detected_image_type=detected_type,
                        is_decorative=is_decorative,
                    )
                )
            elif (
                has_alt
                and existing_alt_text
                and self.validate_alt_text
                and self.image_generator
            ):
                # Image HAS alt text - validate it with AI
                try:
                    image_path = self._extract_image(img)
                    if image_path:
                        img_context = f"Image in Excel sheet '{sheet_name}'"
                        if context.get("headers"):
                            img_context += (
                                f". Sheet headers: {', '.join(context['headers'][:5])}"
                            )

                        # Validate existing alt text with AI
                        validation_result = run_async_from_sync(
                            self.image_generator.validate_alt_text(
                                image_path=image_path,
                                existing_alt_text=existing_alt_text,
                                context=img_context,
                            )
                        )

                        try:
                            os.unlink(image_path)
                        except OSError as e:
                            logger.debug(
                                f"[XlsxProcessor] Could not remove temp image "
                                f"'{image_path}': {e}"
                            )

                        if validation_result.get("success"):
                            is_accurate = validation_result.get("is_accurate", True)
                            accuracy_score = validation_result.get(
                                "accuracy_score", 1.0
                            )
                            validation_issues = validation_result.get("issues", [])
                            suggested_improvement = validation_result.get(
                                "suggested_improvement"
                            )

                            # Only report if alt text is inaccurate or has issues
                            if (
                                not is_accurate
                                or accuracy_score < 0.7
                                or validation_issues
                            ):
                                issues.append(
                                    ImageIssue(
                                        sheet_name=sheet_name,
                                        image_index=idx,
                                        has_alt_text=True,
                                        existing_alt_text=existing_alt_text,
                                        suggested_alt_text=suggested_improvement,
                                        alt_text_validated=True,
                                        alt_text_accurate=is_accurate,
                                        alt_text_issues=validation_issues,
                                        validation_score=accuracy_score,
                                    )
                                )
                except Exception as e:
                    logger.warning(
                        f"[XlsxProcessor] Image alt text validation failed: {e}"
                    )

        return issues

    def _extract_image(self, img: XLImage) -> Optional[str]:
        """Extract image to temp file"""
        try:
            if hasattr(img, "_data") and img._data:
                image = Image.open(BytesIO(img._data()))
                with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as tmp:
                    image.save(tmp, format="PNG")
                    return tmp.name
        except Exception as e:
            logger.warning(f"[XlsxProcessor] Failed to extract image: {e}")
        return None

    def _check_merged_cells(
        self, ws: Worksheet, sheet_name: str, merged_ranges: List
    ) -> List[MergeCellIssue]:
        """Check merged cells for accessibility issues"""
        issues = []

        for merge_range in merged_ranges:
            # Get merge dimensions
            min_col, min_row, max_col, max_row = merge_range.bounds
            rows_merged = max_row - min_row + 1
            cols_merged = max_col - min_col + 1

            # Flag large merges (can confuse screen readers)
            if rows_merged > 3 or cols_merged > 5:
                issues.append(
                    MergeCellIssue(
                        sheet_name=sheet_name,
                        merge_range=str(merge_range),
                        rows_merged=rows_merged,
                        cols_merged=cols_merged,
                        issue_type="large_merge",
                        suggested_fix=f"Large merged cell ({rows_merged}x{cols_merged}) can confuse screen readers. Consider restructuring data.",
                    )
                )

            # Flag merges in first row (likely headers)
            elif min_row == 1 and cols_merged > 1:
                issues.append(
                    MergeCellIssue(
                        sheet_name=sheet_name,
                        merge_range=str(merge_range),
                        rows_merged=rows_merged,
                        cols_merged=cols_merged,
                        issue_type="header_merge",
                        suggested_fix="Merged header cells can break table navigation. Consider using separate cells with clear column headers.",
                    )
                )

        return issues

    def _check_color_coding(
        self, ws: Worksheet, sheet_name: str
    ) -> List[ColorOnlyIssue]:
        """Check for color-only information (no text alternatives)"""
        issues = []

        # Sample cells to check for color patterns
        colors_used = set()
        cells_with_color = []

        # Check first 100 rows max
        max_check_row = min(ws.max_row or 0, 100)
        max_check_col = min(ws.max_column or 0, 20)

        for row in range(1, max_check_row + 1):
            for col in range(1, max_check_col + 1):
                cell = ws.cell(row=row, column=col)

                # Check fill color
                if cell.fill and cell.fill.start_color:
                    rgb = cell.fill.start_color.rgb
                    if rgb and rgb != "00000000" and rgb != "FFFFFFFF":
                        colors_used.add(rgb)
                        cells_with_color.append((row, col, rgb))

                # Check font color (non-black)
                if cell.font and cell.font.color:
                    rgb = cell.font.color.rgb
                    if rgb and rgb != "00000000" and rgb != "FF000000":
                        colors_used.add(rgb)

        # If multiple colors are used extensively, flag potential color-coding
        if len(colors_used) >= 3 and len(cells_with_color) > 10:
            # Check if there's a legend or text explanation
            # (simplified check - look for words like "legend", "key", "color")
            has_legend = False
            for row in range(1, min(ws.max_row or 0, 20) + 1):
                for col in range(1, min(ws.max_column or 0, 10) + 1):
                    cell_value = str(ws.cell(row=row, column=col).value or "").lower()
                    if any(
                        word in cell_value
                        for word in ["legend", "key", "color meaning", "status"]
                    ):
                        has_legend = True
                        break

            if not has_legend:
                issues.append(
                    ColorOnlyIssue(
                        sheet_name=sheet_name,
                        cell_range="Multiple cells",
                        issue_type="color_coding",
                        colors_used=list(colors_used)[:5],  # Limit to 5
                        suggested_fix="Color-coding detected without legend. Add text labels or a legend explaining what colors mean. Color alone should not convey information (WCAG 1.4.1).",
                    )
                )

        return issues

    def _check_navigation(
        self,
        ws: Worksheet,
        sheet_name: str,
        row_count: int,
        col_count: int,
        has_frozen: bool,
    ) -> List[NavigationIssue]:
        """Check navigation aids for large datasets"""
        issues = []

        # Flag large datasets without frozen panes
        if row_count > 20 and col_count > 3 and not has_frozen:
            issues.append(
                NavigationIssue(
                    sheet_name=sheet_name,
                    issue_type="no_freeze",
                    suggested_fix=f"Large dataset ({row_count} rows) without frozen panes. Freeze header row: View > Freeze Panes > Freeze Top Row",
                )
            )

        # Flag very large datasets
        if row_count > 1000:
            issues.append(
                NavigationIssue(
                    sheet_name=sheet_name,
                    issue_type="large_dataset",
                    suggested_fix=f"Very large dataset ({row_count} rows). Consider adding filters, named ranges, or splitting into multiple sheets for better navigation.",
                )
            )

        return issues

    def _analyze_conditional_formatting(
        self, ws: Worksheet, sheet_name: str
    ) -> List[ConditionalFormatIssue]:
        """
        Analyze conditional formatting rules for accessibility issues (WCAG 1.4.1).

        Conditional formatting that uses color alone to convey information
        is problematic for color-blind users and violates WCAG 1.4.1.

        Problematic CF types:
        - colorScale: Uses gradient colors to show values
        - dataBar: Uses colored bars to show values
        - iconSet: Uses icons (usually with colors) to categorize values

        Args:
            ws: The worksheet to check
            sheet_name: Name of the sheet

        Returns:
            List of ConditionalFormatIssue objects for problematic rules
        """
        issues = []

        # Check if worksheet has conditional formatting
        if not hasattr(ws, "conditional_formatting") or not ws.conditional_formatting:
            return issues

        # Iterate through conditional formatting ranges
        for cf_range in ws.conditional_formatting:
            cell_range = str(cf_range)

            # Each cf_range can have multiple rules
            for rule in cf_range.cfRule:
                rule_type = rule.type if hasattr(rule, "type") else "unknown"

                # Check for color-only conditional formatting types
                if rule_type == "colorScale":
                    # Color scales use gradient colors - always color-only
                    color_count = 0
                    if hasattr(rule, "colorScale") and rule.colorScale:
                        color_count = (
                            len(rule.colorScale.color)
                            if hasattr(rule.colorScale, "color")
                            else 2
                        )

                    issues.append(
                        ConditionalFormatIssue(
                            sheet_name=sheet_name,
                            cell_range=cell_range,
                            rule_type="colorScale",
                            uses_color_only=True,
                            has_text_alternative=False,
                            color_count=color_count,
                            recommendation=(
                                "Color scale formatting uses color alone to convey values. "
                                "Add a text column or legend explaining what colors represent. "
                                "Consider using data bars with value labels instead (WCAG 1.4.1)."
                            ),
                        )
                    )

                elif rule_type == "dataBar":
                    # Data bars use colored bars - color-only unless values shown
                    show_value = True  # Default assumption
                    if hasattr(rule, "dataBar") and rule.dataBar:
                        # Check if values are hidden (showValue=False)
                        if hasattr(rule.dataBar, "showValue"):
                            show_value = rule.dataBar.showValue

                    if not show_value:
                        issues.append(
                            ConditionalFormatIssue(
                                sheet_name=sheet_name,
                                cell_range=cell_range,
                                rule_type="dataBar",
                                uses_color_only=True,
                                has_text_alternative=False,
                                color_count=1,
                                recommendation=(
                                    "Data bar formatting hides cell values, using color bars only. "
                                    "Enable 'Show Bar Only' option or ensure cell values are visible. "
                                    "Screen readers cannot interpret visual bar lengths (WCAG 1.4.1)."
                                ),
                            )
                        )

                elif rule_type == "iconSet":
                    # Icon sets use colored icons to categorize
                    show_value = True
                    icon_count = 3  # Default icon set size
                    if hasattr(rule, "iconSet") and rule.iconSet:
                        if hasattr(rule.iconSet, "showValue"):
                            show_value = rule.iconSet.showValue
                        if hasattr(rule.iconSet, "iconSet"):
                            # Count icons based on set type
                            icon_set_type = rule.iconSet.iconSet
                            if "5" in str(icon_set_type):
                                icon_count = 5
                            elif "4" in str(icon_set_type):
                                icon_count = 4

                    if not show_value:
                        issues.append(
                            ConditionalFormatIssue(
                                sheet_name=sheet_name,
                                cell_range=cell_range,
                                rule_type="iconSet",
                                uses_color_only=True,
                                has_text_alternative=False,
                                color_count=icon_count,
                                recommendation=(
                                    "Icon set formatting hides cell values, showing only icons. "
                                    "Enable value display or add a text column with status labels. "
                                    "Icons alone are not accessible to screen readers (WCAG 1.4.1)."
                                ),
                            )
                        )

                elif rule_type == "cellIs":
                    # Cell Is rules (highlight cells) - check if only color is used
                    # These typically apply fill/font color based on conditions
                    has_pattern = False
                    if hasattr(rule, "dxf") and rule.dxf:
                        # Check if differential formatting uses fill
                        if hasattr(rule.dxf, "fill") and rule.dxf.fill:
                            has_pattern = True

                    if has_pattern:
                        # This is common "highlight cells that contain..." formatting
                        # Flag if no text indicator accompanies the color
                        issues.append(
                            ConditionalFormatIssue(
                                sheet_name=sheet_name,
                                cell_range=cell_range,
                                rule_type="cellIs",
                                uses_color_only=True,
                                has_text_alternative=False,
                                color_count=1,
                                recommendation=(
                                    "Cell highlighting uses color to indicate special values. "
                                    "Add a text marker (e.g., '*' or 'Alert') or legend explaining "
                                    "what highlighted cells represent (WCAG 1.4.1)."
                                ),
                            )
                        )

        return issues

    def _analyze_pivot_tables(
        self, ws: Worksheet, sheet_name: str
    ) -> List[PivotTableIssue]:
        """
        Analyze pivot tables for accessibility issues (WCAG 1.3.1).

        Pivot tables have complex structure that can be difficult for
        screen readers to navigate. This method identifies:
        - Complex nested headers
        - Missing field labels
        - Multi-level row/column structures

        Args:
            ws: The worksheet to check
            sheet_name: Name of the sheet

        Returns:
            List of PivotTableIssue objects for problematic pivot tables
        """
        issues = []

        # Check if worksheet has pivot tables
        # In openpyxl, pivot tables are accessed via ws._pivots
        if not hasattr(ws, "_pivots") or not ws._pivots:
            return issues

        for idx, pivot in enumerate(ws._pivots):
            pivot_name = f"PivotTable{idx + 1}"
            pivot_location = "Unknown"
            row_fields = []
            column_fields = []
            value_fields = []
            has_labels = True
            has_grand_totals = False

            try:
                # Get pivot table name if available
                if hasattr(pivot, "name") and pivot.name:
                    pivot_name = pivot.name

                # Get location from pivot table definition
                if hasattr(pivot, "location") and pivot.location:
                    if hasattr(pivot.location, "ref"):
                        pivot_location = pivot.location.ref

                # Analyze pivot table fields
                if hasattr(pivot, "pivotFields") and pivot.pivotFields:
                    for field in pivot.pivotFields:
                        # Check if field has a name/label
                        if hasattr(field, "name") and field.name:
                            # Determine field type based on axis
                            if hasattr(field, "axis"):
                                if field.axis == "axisRow":
                                    row_fields.append(field.name)
                                elif field.axis == "axisCol":
                                    column_fields.append(field.name)

                # Check row fields
                if hasattr(pivot, "rowFields") and pivot.rowFields:
                    for rf in pivot.rowFields:
                        if hasattr(rf, "x") and rf.x is not None:
                            row_fields.append(f"Field{rf.x}")

                # Check column fields
                if hasattr(pivot, "colFields") and pivot.colFields:
                    for cf in pivot.colFields:
                        if hasattr(cf, "x") and cf.x is not None:
                            column_fields.append(f"Field{cf.x}")

                # Check data fields (values)
                if hasattr(pivot, "dataFields") and pivot.dataFields:
                    for df in pivot.dataFields:
                        field_name = (
                            df.name if hasattr(df, "name") and df.name else "Value"
                        )
                        value_fields.append(field_name)

                # Check for grand totals
                if hasattr(pivot, "rowGrandTotals"):
                    has_grand_totals = pivot.rowGrandTotals or False
                if hasattr(pivot, "colGrandTotals"):
                    has_grand_totals = has_grand_totals or (
                        pivot.colGrandTotals or False
                    )

            except Exception as e:
                logger.warning(f"[XlsxProcessor] Error analyzing pivot table: {e}")

            # Determine issue type and generate recommendations
            recommendations = []
            issue_type = "complex_structure"

            # Check for complex nested structure
            total_dimensions = len(row_fields) + len(column_fields)
            if total_dimensions > 2:
                issue_type = "nested_headers"
                recommendations.append(
                    "Pivot table has nested row/column headers which are difficult "
                    "for screen readers to navigate. Consider flattening the structure."
                )

            # Check for missing labels
            if not row_fields and not column_fields:
                issue_type = "missing_labels"
                recommendations.append(
                    "Pivot table field labels could not be detected. "
                    "Ensure all row and column fields have descriptive labels."
                )

            # General recommendations for all pivot tables
            recommendations.append(
                "Consider providing a flat data table as an alternative. "
                "Pivot tables are inherently complex for assistive technologies."
            )

            if not has_grand_totals:
                recommendations.append(
                    "Adding grand totals can help users understand the overall summary."
                )

            recommendations.append(
                "Ensure the source data is accessible and properly structured (WCAG 1.3.1)."
            )

            # Only flag if we found structural issues or if it's a complex pivot
            if total_dimensions >= 1 or issue_type != "complex_structure":
                issues.append(
                    PivotTableIssue(
                        sheet_name=sheet_name,
                        pivot_name=pivot_name,
                        pivot_location=pivot_location,
                        row_fields=row_fields[:5],  # Limit to first 5
                        column_fields=column_fields[:5],
                        value_fields=value_fields[:5],
                        has_field_labels=has_labels,
                        has_grand_totals=has_grand_totals,
                        issue_type=issue_type,
                        recommendations=recommendations[:4],  # Limit recommendations
                    )
                )

        return issues

    def _check_text_contrast(
        self, ws: Worksheet, sheet_name: str
    ) -> List[ContrastIssue]:
        """
        Check text contrast ratios for accessibility (WCAG 1.4.3).

        WCAG 1.4.3 requires:
        - 4.5:1 contrast ratio for normal text
        - 3:1 for large text (18pt or 14pt bold)

        Args:
            ws: The worksheet to check
            sheet_name: Name of the sheet

        Returns:
            List of ContrastIssue objects for cells failing contrast requirements
        """
        issues = []
        WCAG_AA_RATIO = 4.5  # Minimum contrast for normal text

        # Limit checking to first 100 rows and 20 columns for performance
        max_row = min(ws.max_row or 0, 100)
        max_col = min(ws.max_column or 0, 20)

        # Track cells already reported (avoid duplicates)
        reported_cells = set()

        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)

                # Skip empty cells
                if not cell.value:
                    continue

                cell_ref = f"{get_column_letter(col)}{row}"

                # Skip already reported
                if cell_ref in reported_cells:
                    continue

                # Get font color (default black)
                fg_hex = self._get_font_color_hex(cell)

                # Get background color (default white)
                bg_hex = self._get_fill_color_hex(cell)

                # Skip if we couldn't determine colors or it's black on white
                if not fg_hex or not bg_hex:
                    continue
                if fg_hex.lower() == "#000000" and bg_hex.lower() == "#ffffff":
                    continue

                # Calculate contrast ratio
                contrast_ratio = self._calculate_contrast_ratio(fg_hex, bg_hex)

                # Check if it fails WCAG AA
                if contrast_ratio < WCAG_AA_RATIO:
                    reported_cells.add(cell_ref)
                    text_preview = str(cell.value)[:30]
                    if len(str(cell.value)) > 30:
                        text_preview += "..."

                    issues.append(
                        ContrastIssue(
                            sheet_name=sheet_name,
                            cell_reference=cell_ref,
                            text_preview=text_preview,
                            foreground_color=fg_hex,
                            background_color=bg_hex,
                            contrast_ratio=round(contrast_ratio, 2),
                            wcag_aa_pass=False,
                            suggested_fix=(
                                f"Contrast ratio {contrast_ratio:.2f}:1 is below WCAG AA (4.5:1). "
                                f"Darken text color or lighten background color."
                            ),
                        )
                    )

        return issues

    def _get_font_color_hex(self, cell) -> Optional[str]:
        """Get font color as hex string from cell"""
        try:
            if cell.font and cell.font.color:
                color = cell.font.color
                if hasattr(color, "rgb") and color.rgb:
                    rgb = color.rgb
                    if isinstance(rgb, str):
                        if len(rgb) == 8:
                            return f"#{rgb[2:]}"  # Skip alpha
                        elif len(rgb) == 6:
                            return f"#{rgb}"
        except Exception:
            pass
        return "#000000"  # Default black

    def _get_fill_color_hex(self, cell) -> Optional[str]:
        """Get fill/background color as hex string from cell"""
        try:
            if cell.fill and cell.fill.start_color:
                color = cell.fill.start_color
                if hasattr(color, "rgb") and color.rgb:
                    rgb = color.rgb
                    if isinstance(rgb, str):
                        # Skip transparent fills
                        if rgb == "00000000":
                            return "#ffffff"
                        if len(rgb) == 8:
                            return f"#{rgb[2:]}"  # Skip alpha
                        elif len(rgb) == 6:
                            return f"#{rgb}"
        except Exception:
            pass
        return "#ffffff"  # Default white

    def _calculate_contrast_ratio(self, color1: str, color2: str) -> float:
        """
        Calculate contrast ratio between two hex colors (WCAG formula).
        https://www.w3.org/TR/WCAG21/#dfn-contrast-ratio
        """

        def hex_to_rgb(hex_color: str) -> tuple:
            hex_color = hex_color.lstrip("#")
            return tuple(int(hex_color[i : i + 2], 16) for i in (0, 2, 4))

        def relative_luminance(rgb: tuple) -> float:
            def adjust(c):
                c = c / 255.0
                if c <= 0.03928:
                    return c / 12.92
                return ((c + 0.055) / 1.055) ** 2.4

            r, g, b = rgb
            return 0.2126 * adjust(r) + 0.7152 * adjust(g) + 0.0722 * adjust(b)

        try:
            rgb1 = hex_to_rgb(color1)
            rgb2 = hex_to_rgb(color2)

            l1 = relative_luminance(rgb1)
            l2 = relative_luminance(rgb2)

            lighter = max(l1, l2)
            darker = min(l1, l2)

            return (lighter + 0.05) / (darker + 0.05)
        except Exception:
            return 21.0  # Return max contrast if calculation fails

    def _calculate_summary(
        self, sheet_name_issues: List, sheets: List[SheetAnalysis]
    ) -> Dict[str, int]:
        """Calculate summary statistics"""
        total_table_issues = sum(len(s.table_header_issues) for s in sheets)
        total_chart_issues = sum(len(s.chart_issues) for s in sheets)
        total_image_issues = sum(len(s.image_issues) for s in sheets)
        total_merge_issues = sum(len(s.merge_issues) for s in sheets)
        total_color_issues = sum(len(s.color_issues) for s in sheets)
        total_nav_issues = sum(len(s.navigation_issues) for s in sheets)
        total_contrast_issues = sum(len(s.contrast_issues) for s in sheets)
        total_cf_issues = sum(len(s.conditional_format_issues) for s in sheets)
        total_pivot_issues = sum(len(s.pivot_table_issues) for s in sheets)

        return {
            "sheet_name_issues": len(sheet_name_issues),
            "table_header_issues": total_table_issues,
            "chart_issues": total_chart_issues,
            "image_issues": total_image_issues,
            "merge_issues": total_merge_issues,
            "color_issues": total_color_issues,
            "navigation_issues": total_nav_issues,
            "contrast_issues": total_contrast_issues,
            "conditional_format_issues": total_cf_issues,
            "pivot_table_issues": total_pivot_issues,
            "total_issues": (
                len(sheet_name_issues)
                + total_table_issues
                + total_chart_issues
                + total_image_issues
                + total_merge_issues
                + total_color_issues
                + total_nav_issues
                + total_contrast_issues
                + total_cf_issues
                + total_pivot_issues
            ),
        }

    def _calculate_compliance_score(self, summary: Dict, total_elements: int) -> float:
        """Calculate compliance score using unified scoring system"""
        from .compliance_scoring import score_from_severity_counts

        # No early return for total_elements == 0: issues must always count
        # (issue #90 — near-empty workbooks scored 100.0 while listing defects).
        # Map issue types to severity:
        # - Chart/Image alt text = High (WCAG 1.1.1)
        # - Table headers = High (WCAG 1.3.1)
        # - Color-only = Medium (WCAG 1.4.1)
        # - Contrast issues = Medium (WCAG 1.4.3)
        # - Conditional formatting = Medium (WCAG 1.4.1)
        # - Pivot tables = Medium (WCAG 1.3.1)
        # - Sheet names = Low (best practice)
        # - Merge issues = Low (best practice)
        # - Navigation = Low (best practice)

        critical = 0
        high = (
            summary.get("chart_issues", 0)
            + summary.get("image_issues", 0)
            + summary.get("table_header_issues", 0)
        )
        medium = (
            summary.get("color_issues", 0)
            + summary.get("contrast_issues", 0)
            + summary.get("conditional_format_issues", 0)
            + summary.get("pivot_table_issues", 0)
        )
        low = (
            summary.get("sheet_name_issues", 0)
            + summary.get("merge_issues", 0)
            + summary.get("navigation_issues", 0)
        )

        result = score_from_severity_counts(
            critical=critical,
            high=high,
            medium=medium,
            low=low,
            total_elements=total_elements,
        )
        return result.score

    def _generate_remediation_suggestions(self, summary: Dict) -> List[str]:
        """Generate high-level remediation suggestions"""
        suggestions = []

        if summary["sheet_name_issues"] > 0:
            suggestions.append(
                f"Rename {summary['sheet_name_issues']} sheets with generic names. "
                "Use descriptive names like 'Q1 Sales' instead of 'Sheet1'."
            )

        if summary["table_header_issues"] > 0:
            suggestions.append(
                f"Define headers for {summary['table_header_issues']} data tables. "
                "Select data and use Insert > Table, or bold the first row."
            )

        if summary["chart_issues"] > 0:
            suggestions.append(
                f"Add titles/alt text to {summary['chart_issues']} charts. "
                "Click chart > Add Chart Element > Chart Title."
            )

        if summary["image_issues"] > 0:
            suggestions.append(
                f"Add alt text to {summary['image_issues']} images. "
                "Right-click image > Edit Alt Text."
            )

        if summary["merge_issues"] > 0:
            suggestions.append(
                f"Review {summary['merge_issues']} merged cell regions. "
                "Large merges can confuse screen readers."
            )

        if summary["color_issues"] > 0:
            suggestions.append(
                f"Address {summary['color_issues']} color-coding issues. "
                "Add text labels or legend - don't rely on color alone."
            )

        if summary["navigation_issues"] > 0:
            suggestions.append(
                f"Improve navigation for {summary['navigation_issues']} sheets. "
                "Freeze header rows and add filters for large datasets."
            )

        if summary.get("contrast_issues", 0) > 0:
            suggestions.append(
                f"Fix {summary['contrast_issues']} text contrast issues. "
                "Ensure text colors have at least 4.5:1 contrast ratio with backgrounds (WCAG 1.4.3)."
            )

        if summary.get("conditional_format_issues", 0) > 0:
            suggestions.append(
                f"Review {summary['conditional_format_issues']} conditional formatting rules. "
                "Color scales, data bars, and icon sets use color alone to convey information. "
                "Add text indicators or legends to ensure accessibility (WCAG 1.4.1)."
            )

        if summary.get("pivot_table_issues", 0) > 0:
            suggestions.append(
                f"Review {summary['pivot_table_issues']} pivot tables for accessibility. "
                "Pivot tables have complex nested structures that are difficult for screen readers. "
                "Consider providing flat data table alternatives (WCAG 1.3.1)."
            )

        if summary["total_issues"] == 0:
            suggestions.append(
                "No accessibility issues detected. Spreadsheet is accessible!"
            )

        return suggestions

    def _analyze_cvd_accessibility(self, wb) -> List[ColorBlindnessAnalysisResult]:
        """
        Analyze color accessibility for color-blind users.

        Extracts cell colors (fill and font) from Excel workbook
        and tests them against all CVD types.

        Args:
            wb: openpyxl Workbook object

        Returns:
            List of ColorBlindnessAnalysisResult for each unique color pair
        """
        if not self.cvd_simulator:
            return []

        results = []
        color_pairs_seen = set()

        # Helper to convert openpyxl color to hex
        def color_to_hex(color_obj) -> Optional[str]:
            """Convert openpyxl Color to hex string"""
            if color_obj is None:
                return None

            # Check for RGB value
            if hasattr(color_obj, "rgb") and color_obj.rgb:
                rgb = color_obj.rgb
                # RGB might be ARGB (8 chars) or RGB (6 chars)
                if isinstance(rgb, str):
                    if len(rgb) == 8:
                        return f"#{rgb[2:]}"  # Skip alpha
                    elif len(rgb) == 6:
                        return f"#{rgb}"
                return None

            # Check for indexed color (can't convert easily)
            return None

        try:
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                # Sample cells (limit to first 100 rows, 20 columns for performance)
                max_row = min(ws.max_row or 0, 100)
                max_col = min(ws.max_column or 0, 20)

                for row in range(1, max_row + 1):
                    for col in range(1, max_col + 1):
                        cell = ws.cell(row=row, column=col)

                        # Get background color (fill)
                        bg_hex = "#ffffff"  # Default white
                        if cell.fill and cell.fill.start_color:
                            bg_color = color_to_hex(cell.fill.start_color)
                            if bg_color and bg_color.lower() not in [
                                "#000000",
                                "#ffffff",
                            ]:
                                bg_hex = bg_color

                        # Get font color
                        fg_hex = "#000000"  # Default black
                        if cell.font and cell.font.color:
                            font_color = color_to_hex(cell.font.color)
                            if font_color:
                                fg_hex = font_color

                        # Skip default black on white
                        if fg_hex.lower() == "#000000" and bg_hex.lower() == "#ffffff":
                            continue

                        # Skip if already analyzed
                        pair_key = (fg_hex.lower(), bg_hex.lower())
                        if pair_key in color_pairs_seen:
                            continue
                        color_pairs_seen.add(pair_key)

                        # Analyze this color pair
                        try:
                            analysis = self.cvd_simulator.analyze_color_accessibility(
                                foreground=fg_hex, background=bg_hex
                            )
                            if analysis.issues:
                                results.append(analysis)
                        except Exception as e:
                            logger.warning(
                                f"[XlsxProcessor] CVD analysis failed for {fg_hex}/{bg_hex}: {e}"
                            )

        except Exception as e:
            logger.error(f"[XlsxProcessor] CVD analysis failed: {e}")

        return results

    def process_directory(self, directory: str) -> List[XlsxProcessingResult]:
        """
        Batch process all XLSX files in a directory

        Args:
            directory: Path to directory containing XLSX files

        Returns:
            List of XlsxProcessingResult for each file
        """
        results = []
        xlsx_files = list(Path(directory).glob("*.xlsx"))

        for xlsx_file in xlsx_files:
            # Skip temp files
            if xlsx_file.name.startswith("~$"):
                continue

            try:
                logger.info(f"[XlsxProcessor] Processing: {xlsx_file.name}")
                result = self.process_xlsx(str(xlsx_file))
                results.append(result)
            except Exception as e:
                logger.error(f"[XlsxProcessor] Error processing {xlsx_file}: {e}")

        return results


class XlsxBatchProcessor:
    """Batch processor for Excel files"""

    def __init__(self, generate_alt_text: bool = False):
        self.processor = XlsxProcessor(
            generate_chart_descriptions=generate_alt_text,
            generate_alt_text=generate_alt_text,
        )

    def process_directory(self, directory: str) -> List[XlsxProcessingResult]:
        """Process all XLSX files in a directory"""
        return self.processor.process_directory(directory)
