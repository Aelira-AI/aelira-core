"""
Analytics & Issue Tracking API Endpoints

New endpoints for:
- Historical compliance trending (using snapshots)
- Issue tracking and management
- Team collaboration features
- Deadline projection

Author: Aelira Team
Created: November 30, 2025
"""

from fastapi import APIRouter, HTTPException, Depends, Query, status
from pydantic import BaseModel
from typing import List, Optional
from sqlalchemy.orm import Session
from datetime import datetime
import logging

from ..db.database import get_db_dependency
from ..db.models import UserRole
from ..education.snapshot_service import SnapshotService
from ..education.issue_tracking_service import (
    IssueTrackingService,
)

# Setup logging
logger = logging.getLogger(__name__)

# Create router
router = APIRouter(prefix="/analytics", tags=["analytics"])


# Pydantic models for requests/responses
class UpdateIssueStatusRequest(BaseModel):
    status: str  # open, in_progress, resolved, wont_fix, false_positive
    resolution_notes: Optional[str] = None
    resolution_method: Optional[str] = None


class AssignIssueRequest(BaseModel):
    assigned_to: str  # user_id


class AddNoteRequest(BaseModel):
    note: str


class BulkUpdateRequest(BaseModel):
    issue_ids: List[str]
    status: str


from ..auth.dependencies import (  # noqa: E402
    AuthenticatedPrincipal,
    get_authenticated_principal,
    verify_department_access,
)


def _authorize_department_analytics(
    principal: AuthenticatedPrincipal, requested_department_id: str
) -> str:
    """Return canonical tenant scope before any department-wide analytics work."""

    if principal.auth_method == "lti" and not principal.lti_account_wide:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Forbidden")
    verify_department_access(requested_department_id, principal.department_id)
    return principal.department_id


def _require_global_snapshot_authority(principal: AuthenticatedPrincipal) -> None:
    """Restrict all-department snapshot capture to platform operators."""

    if (
        principal.auth_method == "lti"
        or principal.user_role is not UserRole.SUPER_ADMIN
    ):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Forbidden")


def _internal_error(action: str, public_detail: str, error: Exception) -> HTTPException:
    """Log a bounded failure class and return a stable public error."""

    logger.error("%s failed (%s)", action, type(error).__name__)
    return HTTPException(
        status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
        detail=public_detail,
    )


# ==================== Snapshot Endpoints ====================


@router.post("/snapshots/capture/{department_id}")
async def capture_snapshot(
    department_id: str,
    db: Session = Depends(get_db_dependency),
    principal: AuthenticatedPrincipal = Depends(get_authenticated_principal),
):
    """
    Manually trigger a compliance snapshot capture for a department.

    Normally called by a daily cron job, but can be triggered manually.

    Returns:
        Captured snapshot data
    """
    department_id = _authorize_department_analytics(principal, department_id)
    logger.info("Manual compliance snapshot capture requested")

    try:
        snapshot = SnapshotService.capture_daily_snapshot(db, department_id)
        return {
            "success": True,
            "snapshot_id": snapshot.id,
            "snapshot_date": snapshot.snapshot_date.isoformat(),
            "avg_compliance_score": snapshot.avg_compliance_score,
            "total_issues": snapshot.total_issues,
            "days_until_deadline": snapshot.days_until_deadline,
        }
    except ValueError:
        raise HTTPException(status_code=404, detail="Department not found")
    except Exception as e:
        raise _internal_error("Snapshot capture", "Unable to capture snapshot", e)


@router.post("/snapshots/capture-all")
async def capture_all_snapshots(
    db: Session = Depends(get_db_dependency),
    principal: AuthenticatedPrincipal = Depends(get_authenticated_principal),
):
    """
    Capture snapshots for all active departments.

    Should be called by a daily cron job.

    Returns:
        Summary of captured snapshots
    """
    _require_global_snapshot_authority(principal)
    logger.info("Capturing snapshots for all departments")

    try:
        snapshots = SnapshotService.capture_all_departments(db)
        return {
            "success": True,
            "snapshots_captured": len(snapshots),
            "departments": [s.department_id for s in snapshots],
        }
    except Exception as e:
        raise _internal_error(
            "All-department snapshot capture", "Unable to capture snapshots", e
        )


@router.get("/trend/{department_id}")
async def get_historical_trend(
    department_id: str,
    days: int = Query(default=30, ge=7, le=365),
    db: Session = Depends(get_db_dependency),
    principal: AuthenticatedPrincipal = Depends(get_authenticated_principal),
):
    """
    Get historical compliance trend data from snapshots.

    Falls back to computing from scans if no snapshots exist.

    Args:
        days: Number of days to look back (7-365)

    Returns:
        Array of trend data points for charting
    """
    department_id = _authorize_department_analytics(principal, department_id)
    logger.info("Getting %s-day historical trend", days)

    try:
        trend_points = SnapshotService.get_historical_trend(db, department_id, days)

        return {
            "department_id": department_id,
            "period_days": days,
            "data_points": len(trend_points),
            "trend": [
                {
                    "date": p.date,
                    "avg_compliance_score": p.avg_compliance_score,
                    "scan_count": p.scan_count,
                    "total_issues": p.total_issues,
                    "files_compliant": p.files_compliant,
                    "files_needs_work": p.files_needs_work,
                    "files_critical": p.files_critical,
                }
                for p in trend_points
            ],
        }
    except Exception as e:
        raise _internal_error(
            "Historical trend lookup", "Unable to retrieve historical trend", e
        )


@router.get("/trend/{department_id}/analysis")
async def get_trend_analysis(
    department_id: str,
    current_period: int = Query(default=7, ge=1, le=30),
    comparison_period: int = Query(default=7, ge=1, le=30),
    db: Session = Depends(get_db_dependency),
    principal: AuthenticatedPrincipal = Depends(get_authenticated_principal),
):
    """
    Analyze trend comparing current period to previous period.

    Useful for showing week-over-week or month-over-month changes.

    Returns:
        Comparison metrics and trend direction
    """
    department_id = _authorize_department_analytics(principal, department_id)
    logger.info("Getting compliance trend analysis")

    try:
        analysis = SnapshotService.analyze_trend(
            db, department_id, current_period, comparison_period
        )

        return {
            "department_id": department_id,
            "current_period_days": current_period,
            "comparison_period_days": comparison_period,
            "analysis": {
                "current_avg_score": analysis.current_avg_score,
                "previous_avg_score": analysis.previous_avg_score,
                "score_change": analysis.score_change,
                "score_change_pct": analysis.score_change_pct,
                "current_total_issues": analysis.current_total_issues,
                "previous_total_issues": analysis.previous_total_issues,
                "issues_change": analysis.issues_change,
                "issues_change_pct": analysis.issues_change_pct,
                "trend_direction": analysis.trend_direction,
                "on_track_for_deadline": analysis.on_track_for_deadline,
            },
        }
    except Exception as e:
        raise _internal_error("Trend analysis", "Unable to retrieve trend analysis", e)


@router.get("/projection/{department_id}")
async def get_deadline_projection(
    department_id: str,
    db: Session = Depends(get_db_dependency),
    principal: AuthenticatedPrincipal = Depends(get_authenticated_principal),
):
    """
    Get projection for April 2027 deadline compliance.

    Based on historical trend and current improvement rate.

    Returns:
        Projection data including likelihood of meeting deadline
    """
    department_id = _authorize_department_analytics(principal, department_id)
    logger.info("Getting deadline projection")

    try:
        projection = SnapshotService.get_deadline_projection(db, department_id)
        return {"department_id": department_id, "projection": projection}
    except Exception as e:
        raise _internal_error(
            "Deadline projection", "Unable to retrieve deadline projection", e
        )


# ==================== Issue Tracking Endpoints ====================


@router.get("/issues/{department_id}")
async def get_department_issues(
    department_id: str,
    status: Optional[str] = None,
    severity: Optional[str] = None,
    assigned_to: Optional[str] = None,
    limit: int = Query(default=100, le=500),
    offset: int = Query(default=0, ge=0),
    db: Session = Depends(get_db_dependency),
    principal: AuthenticatedPrincipal = Depends(get_authenticated_principal),
):
    """
    Get all tracked issues for a department with optional filters.

    Args:
        status: Filter by status (open, in_progress, resolved, wont_fix, false_positive)
        severity: Filter by severity (critical, high, medium, low)
        assigned_to: Filter by assignee user ID
        limit: Maximum results (default 100)
        offset: Pagination offset

    Returns:
        List of issues with metadata
    """
    department_id = _authorize_department_analytics(principal, department_id)
    logger.info("Getting tracked accessibility issues")

    try:
        issues = IssueTrackingService.get_department_issues(
            db, department_id, status, severity, assigned_to, limit, offset
        )

        return {
            "department_id": department_id,
            "count": len(issues),
            "offset": offset,
            "limit": limit,
            "issues": [
                {
                    "id": i.id,
                    "scan_id": i.scan_id,
                    "file_name": i.file_name,
                    "issue_type": i.issue_type,
                    "severity": i.severity,
                    "wcag_criterion": i.wcag_criterion,
                    "description": i.description,
                    "status": i.status,
                    "assigned_to_name": i.assigned_to_name,
                    "created_at": i.created_at,
                    "updated_at": i.updated_at,
                    "auto_fix_available": i.auto_fix_available,
                    "auto_fix_applied": i.auto_fix_applied,
                }
                for i in issues
            ],
        }
    except Exception as e:
        raise _internal_error("Issue listing", "Unable to retrieve tracked issues", e)


@router.get("/issues/{department_id}/stats")
async def get_issue_stats(
    department_id: str,
    db: Session = Depends(get_db_dependency),
    principal: AuthenticatedPrincipal = Depends(get_authenticated_principal),
):
    """
    Get issue statistics for a department.

    Returns:
        Counts by status and resolution rate
    """
    department_id = _authorize_department_analytics(principal, department_id)
    logger.info("Getting tracked issue statistics")

    try:
        stats = IssueTrackingService.get_issue_stats(db, department_id)

        return {
            "department_id": department_id,
            "stats": {
                "total_issues": stats.total_issues,
                "open_issues": stats.open_issues,
                "in_progress_issues": stats.in_progress_issues,
                "resolved_issues": stats.resolved_issues,
                "wont_fix_issues": stats.wont_fix_issues,
                "false_positive_issues": stats.false_positive_issues,
                "auto_fixable_issues": stats.auto_fixable_issues,
                "auto_fixed_issues": stats.auto_fixed_issues,
                "resolution_rate": stats.resolution_rate,
            },
        }
    except Exception as e:
        raise _internal_error(
            "Issue statistics", "Unable to retrieve issue statistics", e
        )


@router.patch("/issues/{issue_id}/status")
async def update_issue_status(
    issue_id: str,
    request: UpdateIssueStatusRequest,
    db: Session = Depends(get_db_dependency),
    principal: AuthenticatedPrincipal = Depends(get_authenticated_principal),
):
    """
    Update the status of a tracked issue.

    Args:
        status: New status (open, in_progress, resolved, wont_fix, false_positive)
        resolution_notes: Optional notes about the resolution
        resolution_method: Optional method (auto, manual, wont_fix)

    Returns:
        Updated issue
    """
    _authorize_department_analytics(principal, principal.department_id)
    logger.info("Updating tracked issue status")

    try:
        issue = IssueTrackingService.update_issue_status(
            db,
            issue_id,
            principal.department_id,
            request.status,
            principal.user_id,
            request.resolution_notes,
            request.resolution_method,
        )

        return {
            "success": True,
            "issue_id": issue.id,
            "new_status": issue.status.value,
            "updated_at": issue.updated_at.isoformat() if issue.updated_at else None,
        }
    except ValueError:
        raise HTTPException(status_code=404, detail="Issue not found")
    except Exception as e:
        raise _internal_error("Issue status update", "Unable to update issue", e)


@router.post("/issues/{issue_id}/assign")
async def assign_issue(
    issue_id: str,
    request: AssignIssueRequest,
    db: Session = Depends(get_db_dependency),
    principal: AuthenticatedPrincipal = Depends(get_authenticated_principal),
):
    """
    Assign an issue to a team member.

    Automatically sets status to IN_PROGRESS if currently OPEN.

    Returns:
        Updated issue with assignment info
    """
    _authorize_department_analytics(principal, principal.department_id)
    logger.info("Assigning tracked issue")

    try:
        issue = IssueTrackingService.assign_issue(
            db,
            issue_id,
            principal.department_id,
            request.assigned_to,
            principal.user_id,
        )

        return {
            "success": True,
            "issue_id": issue.id,
            "assigned_to": issue.assigned_to,
            "assigned_at": issue.assigned_at.isoformat() if issue.assigned_at else None,
            "status": issue.status.value,
        }
    except ValueError:
        raise HTTPException(status_code=404, detail="Issue or assignee not found")
    except Exception as e:
        raise _internal_error("Issue assignment", "Unable to assign issue", e)


@router.post("/issues/{issue_id}/note")
async def add_issue_note(
    issue_id: str,
    request: AddNoteRequest,
    db: Session = Depends(get_db_dependency),
    principal: AuthenticatedPrincipal = Depends(get_authenticated_principal),
):
    """
    Add a note to an issue for team collaboration.

    Notes are timestamped and attributed to the user.

    Returns:
        Updated issue with new note
    """
    _authorize_department_analytics(principal, principal.department_id)
    logger.info("Adding tracked issue note")

    try:
        issue = IssueTrackingService.add_issue_note(
            db,
            issue_id,
            principal.department_id,
            request.note,
            principal.user_id,
        )

        return {
            "success": True,
            "issue_id": issue.id,
            "notes": issue.notes,
            "updated_at": issue.updated_at.isoformat() if issue.updated_at else None,
        }
    except ValueError:
        raise HTTPException(status_code=404, detail="Issue not found")
    except Exception as e:
        raise _internal_error("Issue note update", "Unable to add issue note", e)


@router.post("/issues/bulk-update")
async def bulk_update_issues(
    request: BulkUpdateRequest,
    db: Session = Depends(get_db_dependency),
    principal: AuthenticatedPrincipal = Depends(get_authenticated_principal),
):
    """
    Bulk update status for multiple issues.

    Useful for mass resolution or triaging.

    Returns:
        Count of updated issues
    """
    _authorize_department_analytics(principal, principal.department_id)
    logger.info("Bulk updating %s tracked issues", len(request.issue_ids))

    try:
        count = IssueTrackingService.bulk_update_status(
            db,
            request.issue_ids,
            principal.department_id,
            request.status,
            principal.user_id,
        )

        return {"success": True, "updated_count": count, "new_status": request.status}
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid issue update request")
    except Exception as e:
        raise _internal_error("Bulk issue update", "Unable to update issues", e)


# ==================== Report & Certificate Endpoints ====================


@router.get("/report/{department_id}")
async def generate_compliance_report(
    department_id: str,
    include_ai_recommendations: bool = Query(default=True),
    db: Session = Depends(get_db_dependency),
    principal: AuthenticatedPrincipal = Depends(get_authenticated_principal),
):
    """
    Generate a comprehensive PDF compliance report for a department.

    Enhanced with:
    - AI-powered recommendations via Gemini (optional)
    - Historical trend analysis
    - Issue tracking status

    Args:
        include_ai_recommendations: Whether to generate AI recommendations (default True)

    Returns:
        PDF file as attachment
    """
    from fastapi.responses import Response
    from ..education.compliance_dashboard import ComplianceDashboard
    from ..education.compliance_report_generator import ComplianceReportGenerator

    department_id = _authorize_department_analytics(principal, department_id)
    logger.info("Generating compliance report")

    try:
        # Get department compliance stats
        stats_obj = ComplianceDashboard.get_department_compliance(db, department_id)
        if not stats_obj:
            raise HTTPException(status_code=404, detail="Department not found")
        stats = stats_obj.to_report_dict()

        # Get trend analysis
        trend_analysis = None
        try:
            analysis = SnapshotService.analyze_trend(db, department_id, 7, 7)
            trend_analysis = {
                "current_avg_score": analysis.current_avg_score,
                "previous_avg_score": analysis.previous_avg_score,
                "score_change": analysis.score_change,
                "score_change_pct": analysis.score_change_pct,
                "current_total_issues": analysis.current_total_issues,
                "previous_total_issues": analysis.previous_total_issues,
                "issues_change": analysis.issues_change,
                "trend_direction": analysis.trend_direction,
                "on_track_for_deadline": analysis.on_track_for_deadline,
            }
        except Exception as e:
            logger.warning("Trend analysis unavailable (%s)", type(e).__name__)

        # Get issue stats
        issue_stats = None
        try:
            stats_obj = IssueTrackingService.get_issue_stats(db, department_id)
            issue_stats = {
                "total_issues": stats_obj.total_issues,
                "open_issues": stats_obj.open_issues,
                "in_progress_issues": stats_obj.in_progress_issues,
                "resolved_issues": stats_obj.resolved_issues,
                "wont_fix_issues": stats_obj.wont_fix_issues,
                "false_positive_issues": stats_obj.false_positive_issues,
                "auto_fixable_issues": stats_obj.auto_fixable_issues,
                "auto_fixed_issues": stats_obj.auto_fixed_issues,
                "resolution_rate": stats_obj.resolution_rate,
            }
        except Exception as e:
            logger.warning("Issue statistics unavailable (%s)", type(e).__name__)

        # Generate AI recommendations if requested
        ai_recommendations = None
        if include_ai_recommendations:
            try:
                ai_recommendations = (
                    await ComplianceReportGenerator.generate_ai_recommendations(
                        stats, trend_analysis, issue_stats
                    )
                )
            except Exception as e:
                logger.warning("AI recommendations unavailable (%s)", type(e).__name__)

        # Generate PDF report
        pdf_bytes = ComplianceReportGenerator.generate_department_report(
            stats=stats,
            trend_analysis=trend_analysis,
            issue_stats=issue_stats,
            ai_recommendations=ai_recommendations,
        )

        # Return as PDF download
        filename = (
            f"compliance_report_{department_id}_{datetime.now().strftime('%Y%m%d')}.pdf"
        )
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename={filename}",
                "Content-Length": str(len(pdf_bytes)),
            },
        )

    except HTTPException:
        raise
    except Exception as e:
        raise _internal_error(
            "Compliance report generation", "Unable to generate compliance report", e
        )


@router.get("/certificate/{department_id}")
async def generate_compliance_certificate(
    department_id: str,
    db: Session = Depends(get_db_dependency),
    principal: AuthenticatedPrincipal = Depends(get_authenticated_principal),
):
    """
    Generate a professional compliance certificate for a department.

    Certificate levels based on compliance score:
    - Platinum (95-100%): Exceptional Compliance Achievement
    - Gold (90-94%): Excellent Compliance Achievement
    - Silver (80-89%): Good Compliance Achievement
    - Bronze (70-79%): Basic Compliance Achievement

    Returns 404 if department score is below 70% (no certificate available).

    Returns:
        PDF certificate as attachment
    """
    from fastapi.responses import Response
    from ..education.compliance_dashboard import ComplianceDashboard
    from ..education.compliance_certificate import ComplianceCertificate

    department_id = _authorize_department_analytics(principal, department_id)
    logger.info("Generating compliance certificate")

    try:
        # Get department compliance stats
        stats = ComplianceDashboard.get_department_compliance(db, department_id)
        if not stats:
            raise HTTPException(status_code=404, detail="Department not found")

        # Get compliance score
        avg_score = stats.avg_compliance_score
        if avg_score < 70:
            raise HTTPException(
                status_code=404,
                detail=f"Department compliance score ({avg_score}/100) is below the minimum threshold (70) for certificate generation",
            )

        # Generate certificate
        pdf_bytes = ComplianceCertificate.generate_certificate(
            department_name=stats.department_name,
            institution=stats.institution,
            compliance_score=avg_score,
            total_scans=stats.total_scans,
            files_analyzed=stats.total_files_scanned,
        )

        if not pdf_bytes:
            raise HTTPException(
                status_code=500, detail="Failed to generate certificate"
            )

        # Determine certificate level for filename
        level = ComplianceCertificate.get_certificate_level(avg_score)
        level_name = level["name"].lower() if level else "certificate"

        filename = f"compliance_certificate_{level_name}_{department_id}_{datetime.now().strftime('%Y%m%d')}.pdf"
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename={filename}",
                "Content-Length": str(len(pdf_bytes)),
            },
        )

    except HTTPException:
        raise
    except Exception as e:
        raise _internal_error(
            "Compliance certificate generation",
            "Unable to generate compliance certificate",
            e,
        )


@router.get("/certificate/{department_id}/eligibility")
async def check_certificate_eligibility(
    department_id: str,
    db: Session = Depends(get_db_dependency),
    principal: AuthenticatedPrincipal = Depends(get_authenticated_principal),
):
    """
    Check if a department is eligible for a compliance certificate.

    Returns:
        Eligibility status and potential certificate level
    """
    from ..education.compliance_dashboard import ComplianceDashboard
    from ..education.compliance_certificate import ComplianceCertificate

    department_id = _authorize_department_analytics(principal, department_id)

    try:
        stats = ComplianceDashboard.get_department_compliance(db, department_id)
        if not stats:
            raise HTTPException(status_code=404, detail="Department not found")

        avg_score = stats.avg_compliance_score
        level = ComplianceCertificate.get_certificate_level(avg_score)

        return {
            "department_id": department_id,
            "compliance_score": avg_score,
            "eligible": level is not None,
            "certificate_level": level["name"] if level else None,
            "description": (
                level["description"] if level else "Score below minimum threshold (70%)"
            ),
            "points_to_next_level": (
                _get_points_to_next_level(avg_score) if level else 70 - avg_score
            ),
        }

    except HTTPException:
        raise
    except Exception as e:
        raise _internal_error(
            "Certificate eligibility", "Unable to check report eligibility", e
        )


def _get_points_to_next_level(score: float) -> float:
    """Calculate points needed to reach next certificate level."""
    if score >= 95:
        return 0  # Already at Platinum
    elif score >= 90:
        return 95 - score  # Points to Platinum
    elif score >= 80:
        return 90 - score  # Points to Gold
    elif score >= 70:
        return 80 - score  # Points to Silver
    else:
        return 70 - score  # Points to Bronze


# ==================== Export Endpoints ====================


@router.get("/export/{department_id}/csv")
async def export_scans_csv(
    department_id: str,
    date_from: Optional[str] = Query(
        default=None, description="Start date (YYYY-MM-DD)"
    ),
    date_to: Optional[str] = Query(default=None, description="End date (YYYY-MM-DD)"),
    db: Session = Depends(get_db_dependency),
    principal: AuthenticatedPrincipal = Depends(get_authenticated_principal),
):
    """
    Export all scan data for a department as CSV.

    Args:
        date_from: Optional start date filter (YYYY-MM-DD)
        date_to: Optional end date filter (YYYY-MM-DD)

    Returns:
        CSV file as attachment
    """
    import csv
    from io import StringIO
    from fastapi.responses import Response
    from ..db.models import Scan, ScanResult

    department_id = _authorize_department_analytics(principal, department_id)
    logger.info("Exporting scan CSV")

    try:
        # Build query
        query = db.query(Scan).filter(Scan.department_id == department_id)

        # Apply date filters
        if date_from:
            try:
                from_date = datetime.strptime(date_from, "%Y-%m-%d")
                query = query.filter(Scan.created_at >= from_date)
            except ValueError:
                raise HTTPException(
                    status_code=400, detail="Invalid date_from format. Use YYYY-MM-DD"
                )

        if date_to:
            try:
                to_date = datetime.strptime(date_to, "%Y-%m-%d")
                # Include the entire end date
                to_date = to_date.replace(hour=23, minute=59, second=59)
                query = query.filter(Scan.created_at <= to_date)
            except ValueError:
                raise HTTPException(
                    status_code=400, detail="Invalid date_to format. Use YYYY-MM-DD"
                )

        scans = query.order_by(Scan.created_at.desc()).all()

        # Create CSV in memory
        output = StringIO()
        writer = csv.writer(output)

        # Header row
        writer.writerow(
            [
                "Scan ID",
                "Created At",
                "File Name",
                "Scan Type",
                "Status",
                "Compliance Score",
                "Critical Issues",
                "High Issues",
                "Medium Issues",
                "Low Issues",
                "Total Issues",
                "Remediated",
            ]
        )

        # Data rows
        for scan in scans:
            # Get scan result if exists
            result = db.query(ScanResult).filter(ScanResult.scan_id == scan.id).first()

            compliance_score = result.compliance_score if result else 0
            critical = result.critical_issues if result else 0
            high = result.high_issues if result else 0
            medium = result.medium_issues if result else 0
            low = result.low_issues if result else 0
            total_issues = critical + high + medium + low

            writer.writerow(
                [
                    scan.id,
                    scan.created_at.isoformat() if scan.created_at else "",
                    scan.file_name or "",
                    scan.scan_type.value if scan.scan_type else "",
                    scan.status.value if scan.status else "",
                    f"{compliance_score:.1f}",
                    critical,
                    high,
                    medium,
                    low,
                    total_issues,
                    "Yes" if getattr(scan, "remediated", False) else "No",
                ]
            )

        csv_content = output.getvalue()
        output.close()

        filename = (
            f"scans_export_{department_id}_{datetime.now().strftime('%Y%m%d')}.csv"
        )
        return Response(
            content=csv_content,
            media_type="text/csv",
            headers={
                "Content-Disposition": f"attachment; filename={filename}",
                "Content-Length": str(len(csv_content)),
            },
        )

    except HTTPException:
        raise
    except Exception as e:
        raise _internal_error("CSV export", "Unable to export scan data", e)


@router.get("/export/{department_id}/excel")
async def export_scans_excel(
    department_id: str,
    date_from: Optional[str] = Query(
        default=None, description="Start date (YYYY-MM-DD)"
    ),
    date_to: Optional[str] = Query(default=None, description="End date (YYYY-MM-DD)"),
    db: Session = Depends(get_db_dependency),
    principal: AuthenticatedPrincipal = Depends(get_authenticated_principal),
):
    """
    Export all scan data for a department as Excel (.xlsx).

    Includes multiple sheets:
    - Summary: Department overview statistics
    - All Scans: Detailed scan data
    - Issues Breakdown: Issues by type and severity

    Args:
        date_from: Optional start date filter (YYYY-MM-DD)
        date_to: Optional end date filter (YYYY-MM-DD)

    Returns:
        Excel file as attachment
    """
    from io import BytesIO
    from fastapi.responses import Response
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from ..db.models import Scan, ScanResult
    from ..education.compliance_dashboard import ComplianceDashboard

    department_id = _authorize_department_analytics(principal, department_id)
    logger.info("Exporting scan workbook")

    try:
        # Build query
        query = db.query(Scan).filter(Scan.department_id == department_id)

        # Apply date filters
        if date_from:
            try:
                from_date = datetime.strptime(date_from, "%Y-%m-%d")
                query = query.filter(Scan.created_at >= from_date)
            except ValueError:
                raise HTTPException(
                    status_code=400, detail="Invalid date_from format. Use YYYY-MM-DD"
                )

        if date_to:
            try:
                to_date = datetime.strptime(date_to, "%Y-%m-%d")
                to_date = to_date.replace(hour=23, minute=59, second=59)
                query = query.filter(Scan.created_at <= to_date)
            except ValueError:
                raise HTTPException(
                    status_code=400, detail="Invalid date_to format. Use YYYY-MM-DD"
                )

        scans = query.order_by(Scan.created_at.desc()).all()

        # Get department stats
        stats = ComplianceDashboard.get_department_compliance(db, department_id)

        # Create workbook
        wb = Workbook()

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="4F46E5", end_color="4F46E5", fill_type="solid"
        )
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Red/Yellow/Green fills for compliance scores
        red_fill = PatternFill(
            start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"
        )
        yellow_fill = PatternFill(
            start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"
        )
        green_fill = PatternFill(
            start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"
        )

        # ==================== Summary Sheet ====================
        ws_summary = wb.active
        ws_summary.title = "Summary"

        summary_data = [
            ["Department Compliance Summary"],
            [""],
            ["Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["Department ID", department_id],
            [""],
            ["Overview"],
            [
                "Total Scans",
                stats.total_scans if stats else len(scans),
            ],
            [
                "Total Files Scanned",
                stats.total_files_scanned if stats else 0,
            ],
            [
                "Average Compliance Score",
                f"{stats.avg_compliance_score:.1f}%" if stats else "N/A",
            ],
            [""],
            ["Issues Summary"],
            [
                "Critical Issues",
                stats.total_critical if stats else 0,
            ],
            [
                "High Issues",
                stats.total_high if stats else 0,
            ],
            [
                "Medium Issues",
                stats.total_medium if stats else 0,
            ],
            [
                "Low Issues",
                stats.total_low if stats else 0,
            ],
        ]

        for row_idx, row in enumerate(summary_data, 1):
            for col_idx, value in enumerate(row, 1):
                cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:
                    cell.font = Font(bold=True, size=14)
                elif row_idx in [6, 11]:
                    cell.font = Font(bold=True)

        ws_summary.column_dimensions["A"].width = 25
        ws_summary.column_dimensions["B"].width = 30

        # ==================== All Scans Sheet ====================
        ws_scans = wb.create_sheet("All Scans")

        headers = [
            "Scan ID",
            "Created At",
            "File Name",
            "Scan Type",
            "Status",
            "Compliance Score",
            "Critical",
            "High",
            "Medium",
            "Low",
            "Total Issues",
            "Remediated",
        ]

        for col_idx, header in enumerate(headers, 1):
            cell = ws_scans.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        for row_idx, scan in enumerate(scans, 2):
            result = db.query(ScanResult).filter(ScanResult.scan_id == scan.id).first()

            compliance_score = result.compliance_score if result else 0
            critical = result.critical_issues if result else 0
            high = result.high_issues if result else 0
            medium = result.medium_issues if result else 0
            low = result.low_issues if result else 0
            total_issues = critical + high + medium + low

            row_data = [
                scan.id,
                scan.created_at.strftime("%Y-%m-%d %H:%M") if scan.created_at else "",
                scan.file_name or "",
                scan.scan_type.value if scan.scan_type else "",
                scan.status.value if scan.status else "",
                compliance_score,
                critical,
                high,
                medium,
                low,
                total_issues,
                "Yes" if getattr(scan, "remediated", False) else "No",
            ]

            for col_idx, value in enumerate(row_data, 1):
                cell = ws_scans.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border

                # Color code compliance score
                if col_idx == 6:  # Compliance Score column
                    if compliance_score >= 90:
                        cell.fill = green_fill
                    elif compliance_score >= 70:
                        cell.fill = yellow_fill
                    else:
                        cell.fill = red_fill

        # Auto-adjust column widths
        for col_idx, header in enumerate(headers, 1):
            column_letter = get_column_letter(col_idx)
            ws_scans.column_dimensions[column_letter].width = max(len(header) + 2, 12)

        # ==================== Issues Breakdown Sheet ====================
        ws_issues = wb.create_sheet("Issues Breakdown")

        # Get issue stats
        issue_stats = None
        try:
            issue_stats = IssueTrackingService.get_issue_stats(db, department_id)
        except Exception:
            pass

        issue_headers = ["Metric", "Count"]
        for col_idx, header in enumerate(issue_headers, 1):
            cell = ws_issues.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        if issue_stats:
            issue_data = [
                ["Total Issues", issue_stats.total_issues],
                ["Open", issue_stats.open_issues],
                ["In Progress", issue_stats.in_progress_issues],
                ["Resolved", issue_stats.resolved_issues],
                ["Won't Fix", issue_stats.wont_fix_issues],
                ["False Positive", issue_stats.false_positive_issues],
                ["", ""],
                ["Auto-Fixable", issue_stats.auto_fixable_issues],
                ["Auto-Fixed", issue_stats.auto_fixed_issues],
                ["Resolution Rate", f"{issue_stats.resolution_rate:.1f}%"],
            ]
        else:
            issue_data = [["No issue tracking data available", ""]]

        for row_idx, row in enumerate(issue_data, 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws_issues.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border

        ws_issues.column_dimensions["A"].width = 20
        ws_issues.column_dimensions["B"].width = 15

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        excel_content = output.getvalue()
        output.close()

        filename = (
            f"scans_export_{department_id}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        )
        return Response(
            content=excel_content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}",
                "Content-Length": str(len(excel_content)),
            },
        )

    except HTTPException:
        raise
    except Exception as e:
        raise _internal_error("Workbook export", "Unable to export scan data", e)


@router.get("/export/{department_id}/bulk")
async def export_bulk_zip(
    department_id: str,
    include_pdfs: bool = Query(
        default=False, description="Include individual scan PDFs"
    ),
    include_certificate: bool = Query(
        default=True, description="Include certificate if eligible"
    ),
    date_from: Optional[str] = Query(
        default=None, description="Start date (YYYY-MM-DD)"
    ),
    date_to: Optional[str] = Query(default=None, description="End date (YYYY-MM-DD)"),
    db: Session = Depends(get_db_dependency),
    principal: AuthenticatedPrincipal = Depends(get_authenticated_principal),
):
    """
    Export all department data as a ZIP file.

    Contents:
    - summary.csv: All scan data
    - compliance_report.pdf: Department compliance report
    - certificate.pdf: Compliance certificate (if eligible and requested)
    - individual_reports/: Per-scan PDF reports (if requested)

    Args:
        include_pdfs: Include individual scan PDF reports (default False)
        include_certificate: Include certificate if eligible (default True)
        date_from: Optional start date filter (YYYY-MM-DD)
        date_to: Optional end date filter (YYYY-MM-DD)

    Returns:
        ZIP file as attachment
    """
    import zipfile
    import csv
    from io import BytesIO, StringIO
    from fastapi.responses import Response
    from ..db.models import Scan, ScanResult
    from ..education.compliance_dashboard import ComplianceDashboard
    from ..education.compliance_report_generator import ComplianceReportGenerator
    from ..education.compliance_certificate import ComplianceCertificate

    department_id = _authorize_department_analytics(principal, department_id)
    logger.info("Exporting bulk analytics archive")

    try:
        # Create ZIP in memory
        zip_buffer = BytesIO()

        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
            # Build query
            query = db.query(Scan).filter(Scan.department_id == department_id)

            if date_from:
                try:
                    from_date = datetime.strptime(date_from, "%Y-%m-%d")
                    query = query.filter(Scan.created_at >= from_date)
                except ValueError:
                    raise HTTPException(
                        status_code=400, detail="Invalid date_from format"
                    )

            if date_to:
                try:
                    to_date = datetime.strptime(date_to, "%Y-%m-%d")
                    to_date = to_date.replace(hour=23, minute=59, second=59)
                    query = query.filter(Scan.created_at <= to_date)
                except ValueError:
                    raise HTTPException(
                        status_code=400, detail="Invalid date_to format"
                    )

            scans = query.order_by(Scan.created_at.desc()).all()

            # 1. Add summary.csv
            csv_output = StringIO()
            writer = csv.writer(csv_output)
            writer.writerow(
                [
                    "Scan ID",
                    "Created At",
                    "File Name",
                    "Scan Type",
                    "Status",
                    "Compliance Score",
                    "Critical",
                    "High",
                    "Medium",
                    "Low",
                    "Total Issues",
                ]
            )

            for scan in scans:
                result = (
                    db.query(ScanResult).filter(ScanResult.scan_id == scan.id).first()
                )
                compliance_score = result.compliance_score if result else 0
                critical = result.critical_issues if result else 0
                high = result.high_issues if result else 0
                medium = result.medium_issues if result else 0
                low = result.low_issues if result else 0

                writer.writerow(
                    [
                        scan.id,
                        scan.created_at.isoformat() if scan.created_at else "",
                        scan.file_name or "",
                        scan.scan_type.value if scan.scan_type else "",
                        scan.status.value if scan.status else "",
                        f"{compliance_score:.1f}",
                        critical,
                        high,
                        medium,
                        low,
                        critical + high + medium + low,
                    ]
                )

            zip_file.writestr("summary.csv", csv_output.getvalue())

            # 2. Add compliance_report.pdf
            try:
                stats = ComplianceDashboard.get_department_compliance(db, department_id)
                if stats:
                    stats_dict = stats.to_report_dict()
                    # Get trend analysis
                    trend_analysis = None
                    try:
                        analysis = SnapshotService.analyze_trend(
                            db, department_id, 7, 7
                        )
                        trend_analysis = {
                            "current_avg_score": analysis.current_avg_score,
                            "previous_avg_score": analysis.previous_avg_score,
                            "score_change": analysis.score_change,
                            "trend_direction": analysis.trend_direction,
                            "on_track_for_deadline": analysis.on_track_for_deadline,
                        }
                    except Exception:
                        pass

                    # Get issue stats
                    issue_stats = None
                    try:
                        stats_obj = IssueTrackingService.get_issue_stats(
                            db, department_id
                        )
                        issue_stats = {
                            "total_issues": stats_obj.total_issues,
                            "open_issues": stats_obj.open_issues,
                            "resolved_issues": stats_obj.resolved_issues,
                            "resolution_rate": stats_obj.resolution_rate,
                        }
                    except Exception:
                        pass

                    pdf_bytes = ComplianceReportGenerator.generate_department_report(
                        stats=stats_dict,
                        trend_analysis=trend_analysis,
                        issue_stats=issue_stats,
                        ai_recommendations=None,  # Skip AI for bulk export speed
                    )
                    zip_file.writestr("compliance_report.pdf", pdf_bytes)
            except Exception as e:
                logger.warning(
                    "Bulk compliance report unavailable (%s)", type(e).__name__
                )

            # 3. Add certificate.pdf (if eligible and requested)
            if include_certificate:
                try:
                    if stats:
                        avg_score = stats.avg_compliance_score
                        if avg_score >= 70:
                            cert_bytes = ComplianceCertificate.generate_certificate(
                                department_name=stats.department_name,
                                institution=stats.institution,
                                compliance_score=avg_score,
                                total_scans=stats.total_scans,
                                files_analyzed=stats.total_files_scanned,
                            )
                            if cert_bytes:
                                level = ComplianceCertificate.get_certificate_level(
                                    avg_score
                                )
                                level_name = (
                                    level["name"].lower() if level else "certificate"
                                )
                                zip_file.writestr(
                                    f"certificate_{level_name}.pdf", cert_bytes
                                )
                except Exception as e:
                    logger.warning(
                        "Bulk certificate output unavailable (%s)", type(e).__name__
                    )

            # 4. Add individual scan PDFs (if requested)
            if include_pdfs and scans:
                from ..education.pdf_report_generator import (
                    AccessibilityPDFReportGenerator,
                )

                for scan in scans[:50]:  # Limit to 50 to avoid timeout
                    try:
                        result = (
                            db.query(ScanResult)
                            .filter(ScanResult.scan_id == scan.id)
                            .first()
                        )
                        if result and result.issues:
                            report_bytes = (
                                AccessibilityPDFReportGenerator.generate_report(
                                    scan_id=scan.id,
                                    file_name=scan.file_name or "Unknown",
                                    scan_type=(
                                        scan.scan_type.value
                                        if scan.scan_type
                                        else "unknown"
                                    ),
                                    compliance_score=result.compliance_score or 0,
                                    issues=result.issues or [],
                                    suggestions=result.suggestions or [],
                                )
                            )
                            safe_filename = (
                                (scan.file_name or scan.id)
                                .replace("/", "_")
                                .replace("\\", "_")
                            )
                            zip_file.writestr(
                                f"individual_reports/{safe_filename}.pdf", report_bytes
                            )
                    except Exception as e:
                        logger.warning(
                            "Individual scan report unavailable (%s)",
                            type(e).__name__,
                        )

            # 5. Add README.txt
            readme_content = f"""Aelira Accessibility Compliance Export
======================================

Department ID: {department_id}
Export Date: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
Total Scans: {len(scans)}

Contents:
- summary.csv: All scan data in CSV format
- compliance_report.pdf: Department compliance report
{"- certificate_*.pdf: Compliance certificate" if include_certificate else ""}
{"- individual_reports/: Per-scan PDF reports" if include_pdfs else ""}

Generated by Aelira - https://example.com
"""
            zip_file.writestr("README.txt", readme_content)

        zip_content = zip_buffer.getvalue()
        zip_buffer.close()

        filename = (
            f"aelira_export_{department_id}_{datetime.now().strftime('%Y%m%d')}.zip"
        )
        return Response(
            content=zip_content,
            media_type="application/zip",
            headers={
                "Content-Disposition": f"attachment; filename={filename}",
                "Content-Length": str(len(zip_content)),
            },
        )

    except HTTPException:
        raise
    except Exception as e:
        raise _internal_error("Bulk export", "Unable to export analytics archive", e)


# ==================== ML-based Compliance Prediction ====================


@router.get("/predict/{department_id}")
async def predict_deadline_compliance(
    department_id: str,
    db: Session = Depends(get_db_dependency),
    principal: AuthenticatedPrincipal = Depends(get_authenticated_principal),
):
    """
    ML-based prediction of April 2027 deadline compliance

     AI-POWERED COMPLIANCE PREDICTION
    REQUIRES API KEY IN PRODUCTION

    Uses multiple predictive models in an ensemble approach to forecast
    whether the department will meet the April 26, 2027 WCAG compliance deadline.

    Models used:
    - Linear trend extrapolation
    - Exponential smoothing with acceleration
    - Logistic regression probability
    - Engagement-based prediction

    Returns:
    - will_meet_deadline: Boolean prediction
    - probability: 0-1 probability of meeting deadline
    - confidence: 0-1 confidence in the prediction
    - projected_score: Expected compliance score at deadline
    - risk_assessment: Level and specific risk factors
    - recommendations: Prioritized action items

    Note: Requires at least 7 days of scanning history for accurate predictions.
    """
    department_id = _authorize_department_analytics(principal, department_id)

    try:
        from ..education.compliance_predictor import predict_compliance

        result = predict_compliance(db, department_id)
        return result

    except Exception as e:
        raise _internal_error(
            "Compliance prediction", "Unable to generate compliance prediction", e
        )


# ==================== Alt Text Quality Analytics ====================


@router.get("/alt-text-quality/{department_id}")
async def get_alt_text_quality_metrics(
    department_id: str,
    days: int = Query(default=30, ge=1, le=365),
    db: Session = Depends(get_db_dependency),
    principal: AuthenticatedPrincipal = Depends(get_authenticated_principal),
):
    """
    Get aggregate alt text quality metrics for a department

     ALT TEXT QUALITY ANALYTICS
    REQUIRES API KEY IN PRODUCTION

    Aggregates alt text quality data from scans to provide department-wide
    quality metrics for reporting and tracking improvements over time.

    Returns:
    - overall_average_score: Department-wide average quality score (0-100)
    - average_grade: A/B/C/D/F based on average
    - grade_distribution: Count of images at each grade level
    - wcag_compliance_rate: Percentage meeting WCAG 2.1 AA
    - common_issues: Most frequently occurring alt text problems
    - trend: Quality score changes over the time period
    - improvement_opportunities: Specific recommendations

    Query params:
    - days: Number of days to analyze (default 30, max 365)
    """
    department_id = _authorize_department_analytics(principal, department_id)

    try:
        from ..db.models import Scan, ScanResult

        # Get scans from the time period
        from datetime import timedelta

        start_date = datetime.utcnow() - timedelta(days=days)

        scans = (
            db.query(Scan)
            .filter(
                Scan.department_id == department_id,
                Scan.created_at >= start_date,
            )
            .all()
        )

        if not scans:
            return {
                "success": True,
                "department_id": department_id,
                "period_days": days,
                "message": "No scans found in the specified period",
                "overall_average_score": None,
                "total_images_analyzed": 0,
            }

        # Aggregate alt text quality data from scan results
        total_score = 0
        total_images = 0
        grade_counts = {"A": 0, "B": 0, "C": 0, "D": 0, "F": 0}
        wcag_pass_count = 0
        all_issues = []

        for scan in scans:
            result = db.query(ScanResult).filter(ScanResult.scan_id == scan.id).first()
            if result and result.issues:
                for issue in result.issues:
                    # Look for image-related issues
                    if "image" in str(issue).lower() or "alt" in str(issue).lower():
                        total_images += 1

                        # Extract score if available in issue data
                        if isinstance(issue, dict):
                            score = issue.get("alt_text_quality_score", 70)
                            total_score += score

                            # Calculate grade
                            if score >= 90:
                                grade_counts["A"] += 1
                            elif score >= 80:
                                grade_counts["B"] += 1
                            elif score >= 70:
                                grade_counts["C"] += 1
                                wcag_pass_count += 1
                            elif score >= 60:
                                grade_counts["D"] += 1
                            else:
                                grade_counts["F"] += 1

                            # Collect issues
                            if issue.get("alt_text_issues"):
                                all_issues.extend(issue.get("alt_text_issues", []))

        # Calculate averages
        avg_score = total_score / total_images if total_images > 0 else 0

        # Determine average grade
        if avg_score >= 90:
            avg_grade = "A"
        elif avg_score >= 80:
            avg_grade = "B"
        elif avg_score >= 70:
            avg_grade = "C"
        elif avg_score >= 60:
            avg_grade = "D"
        else:
            avg_grade = "F"

        # Count common issues
        issue_counts = {}
        for issue in all_issues:
            issue_str = str(issue) if not isinstance(issue, str) else issue
            issue_counts[issue_str] = issue_counts.get(issue_str, 0) + 1
        common_issues = sorted(issue_counts.items(), key=lambda x: -x[1])[:10]

        # Calculate WCAG compliance rate
        wcag_rate = (wcag_pass_count / total_images * 100) if total_images > 0 else 0

        return {
            "success": True,
            "department_id": department_id,
            "period_days": days,
            "total_scans_analyzed": len(scans),
            "total_images_analyzed": total_images,
            "overall_average_score": round(avg_score, 1),
            "average_grade": avg_grade,
            "grade_distribution": grade_counts,
            "wcag_compliance_rate": round(wcag_rate, 1),
            "common_issues": [
                {"issue": issue, "count": count} for issue, count in common_issues
            ],
            "improvement_opportunities": [
                {
                    "priority": "high",
                    "recommendation": "Focus on images with missing alt text first",
                    "impact": "Fixes WCAG 1.1.1 violations",
                },
                {
                    "priority": "medium",
                    "recommendation": "Remove redundant 'image of' prefixes from alt text",
                    "impact": "Improves screen reader experience",
                },
                {
                    "priority": "low",
                    "recommendation": "Add context-specific descriptions for complex images",
                    "impact": "Better conveys educational content",
                },
            ],
            "note": "For detailed per-image scoring, use POST /education/image/score-alt-text",
        }

    except Exception as e:
        raise _internal_error(
            "Alt-text quality analytics",
            "Unable to retrieve alt-text quality metrics",
            e,
        )


# Alias for education router compatibility
education_router = router
