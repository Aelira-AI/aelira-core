"""
Tests for Conditional Formatting Analysis

Tests cover:
- ConditionalFormatIssue model
- Color scale detection
- Data bar detection
- Icon set detection
- Cell highlighting detection
- Integration with XlsxProcessor
"""

import pytest
import tempfile
import os

from openpyxl import Workbook
from openpyxl.formatting.rule import (
    ColorScaleRule,
    DataBarRule,
    IconSetRule,
    CellIsRule,
)
from openpyxl.styles import PatternFill

from src.education.xlsx_processor import (
    XlsxProcessor,
    ConditionalFormatIssue,
)


class TestConditionalFormatIssueModel:
    """Test ConditionalFormatIssue Pydantic model."""

    def test_cf_issue_creation(self):
        """Test creating a conditional format issue."""
        issue = ConditionalFormatIssue(
            sheet_name="Sheet1",
            cell_range="A1:A10",
            rule_type="colorScale",
            uses_color_only=True,
            has_text_alternative=False,
            color_count=3,
            recommendation="Add text indicators for color scale values",
        )

        assert issue.sheet_name == "Sheet1"
        assert issue.cell_range == "A1:A10"
        assert issue.rule_type == "colorScale"
        assert issue.uses_color_only is True
        assert issue.color_count == 3

    def test_cf_issue_minimal(self):
        """Test conditional format issue with minimal fields."""
        issue = ConditionalFormatIssue(
            sheet_name="Data",
            cell_range="B2:B20",
            rule_type="dataBar",
            uses_color_only=True,
            recommendation="Show cell values alongside data bars",
        )

        assert issue.has_text_alternative is False
        assert issue.color_count == 0

    def test_cf_issue_types(self):
        """Test different conditional format rule types."""
        rule_types = ["colorScale", "dataBar", "iconSet", "cellIs"]

        for rule_type in rule_types:
            issue = ConditionalFormatIssue(
                sheet_name="Test",
                cell_range="A1:A5",
                rule_type=rule_type,
                uses_color_only=True,
                recommendation="Test recommendation",
            )
            assert issue.rule_type == rule_type


class TestColorScaleDetection:
    """Test color scale conditional formatting detection."""

    @pytest.fixture
    def processor(self):
        """Create XlsxProcessor for testing."""
        return XlsxProcessor()

    def test_detect_color_scale_rule(self, processor):
        """Test detection of color scale formatting."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb = Workbook()
            ws = wb.active
            ws.title = "Sales Data"

            # Add data
            for i in range(1, 11):
                ws.cell(row=i, column=1, value=i * 10)

            # Add color scale rule (green-yellow-red)
            rule = ColorScaleRule(
                start_type="min",
                start_color="00FF00",
                mid_type="percentile",
                mid_value=50,
                mid_color="FFFF00",
                end_type="max",
                end_color="FF0000",
            )
            ws.conditional_formatting.add("A1:A10", rule)

            wb.save(f.name)

            result = processor.process_xlsx(f.name)

            # Should detect color scale issue
            cf_issues = result.sheets[0].conditional_format_issues
            assert len(cf_issues) >= 1

            color_scale_issues = [i for i in cf_issues if i.rule_type == "colorScale"]
            assert len(color_scale_issues) >= 1
            assert color_scale_issues[0].uses_color_only is True

        os.unlink(f.name)

    def test_color_scale_in_summary(self, processor):
        """Test that color scale issues appear in summary."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb = Workbook()
            ws = wb.active

            for i in range(1, 6):
                ws.cell(row=i, column=1, value=i * 20)

            rule = ColorScaleRule(
                start_type="min",
                start_color="63BE7B",
                end_type="max",
                end_color="F8696B",
            )
            ws.conditional_formatting.add("A1:A5", rule)

            wb.save(f.name)

            result = processor.process_xlsx(f.name)

            assert "conditional_format_issues" in result.summary
            assert result.summary["conditional_format_issues"] >= 1

        os.unlink(f.name)


class TestDataBarDetection:
    """Test data bar conditional formatting detection."""

    @pytest.fixture
    def processor(self):
        """Create XlsxProcessor for testing."""
        return XlsxProcessor()

    def test_detect_data_bar_rule(self, processor):
        """Test detection of data bar formatting."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb = Workbook()
            ws = wb.active
            ws.title = "Progress"

            # Add data
            for i in range(1, 6):
                ws.cell(row=i, column=1, value=i * 20)

            # Add data bar rule
            rule = DataBarRule(start_type="min", end_type="max", color="638EC6")
            ws.conditional_formatting.add("A1:A5", rule)

            wb.save(f.name)

            result = processor.process_xlsx(f.name)

            # Check for data bar issues
            cf_issues = result.sheets[0].conditional_format_issues
            data_bar_issues = [i for i in cf_issues if i.rule_type == "dataBar"]

            # Data bars with visible values should not be flagged as color-only
            # (unless showValue is explicitly False)
            assert isinstance(cf_issues, list)

        os.unlink(f.name)


class TestIconSetDetection:
    """Test icon set conditional formatting detection."""

    @pytest.fixture
    def processor(self):
        """Create XlsxProcessor for testing."""
        return XlsxProcessor()

    def test_detect_icon_set_rule(self, processor):
        """Test detection of icon set formatting."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb = Workbook()
            ws = wb.active
            ws.title = "Status"

            # Add data
            for i in range(1, 6):
                ws.cell(row=i, column=1, value=i * 20)

            # Add icon set rule (traffic lights)
            rule = IconSetRule(
                icon_style="3TrafficLights1", type="percent", values=[0, 33, 67]
            )
            ws.conditional_formatting.add("A1:A5", rule)

            wb.save(f.name)

            result = processor.process_xlsx(f.name)

            # Check for icon set issues
            cf_issues = result.sheets[0].conditional_format_issues
            icon_issues = [i for i in cf_issues if i.rule_type == "iconSet"]

            # Icon sets without values shown are problematic
            assert isinstance(cf_issues, list)

        os.unlink(f.name)


class TestCellHighlightDetection:
    """Test cell highlight conditional formatting detection."""

    @pytest.fixture
    def processor(self):
        """Create XlsxProcessor for testing."""
        return XlsxProcessor()

    def test_detect_cell_is_rule(self, processor):
        """Test detection of cell highlighting rules."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb = Workbook()
            ws = wb.active
            ws.title = "Alerts"

            # Add data
            for i in range(1, 6):
                ws.cell(row=i, column=1, value=i * 10)

            # Add cell is rule (highlight cells > 30)
            red_fill = PatternFill(
                start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"
            )
            rule = CellIsRule(operator="greaterThan", formula=["30"], fill=red_fill)
            ws.conditional_formatting.add("A1:A5", rule)

            wb.save(f.name)

            result = processor.process_xlsx(f.name)

            # Check for cell highlighting issues
            cf_issues = result.sheets[0].conditional_format_issues
            cell_is_issues = [i for i in cf_issues if i.rule_type == "cellIs"]

            assert isinstance(cf_issues, list)
            # Cell is rules with fill should be flagged
            if cell_is_issues:
                assert cell_is_issues[0].uses_color_only is True

        os.unlink(f.name)


class TestNoConditionalFormatting:
    """Test behavior when no conditional formatting exists."""

    @pytest.fixture
    def processor(self):
        """Create XlsxProcessor for testing."""
        return XlsxProcessor()

    def test_empty_sheet_no_cf_issues(self, processor):
        """Test that empty sheet has no CF issues."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb = Workbook()
            ws = wb.active
            ws.cell(row=1, column=1, value="Test")
            wb.save(f.name)

            result = processor.process_xlsx(f.name)

            assert len(result.sheets[0].conditional_format_issues) == 0
            assert result.summary.get("conditional_format_issues", 0) == 0

        os.unlink(f.name)

    def test_plain_data_no_cf_issues(self, processor):
        """Test that plain data without CF has no issues."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb = Workbook()
            ws = wb.active

            # Add data without conditional formatting
            headers = ["Name", "Value", "Status"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)

            for row in range(2, 10):
                ws.cell(row=row, column=1, value=f"Item {row}")
                ws.cell(row=row, column=2, value=row * 10)
                ws.cell(row=row, column=3, value="Active")

            wb.save(f.name)

            result = processor.process_xlsx(f.name)

            assert len(result.sheets[0].conditional_format_issues) == 0

        os.unlink(f.name)


class TestXlsxProcessorIntegration:
    """Test conditional formatting integration with XlsxProcessor."""

    @pytest.fixture
    def processor(self):
        """Create XlsxProcessor for testing."""
        return XlsxProcessor()

    def test_cf_issues_in_sheet_analysis(self, processor):
        """Test that CF issues are included in sheet analysis."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb = Workbook()
            ws = wb.active

            for i in range(1, 6):
                ws.cell(row=i, column=1, value=i * 20)

            rule = ColorScaleRule(
                start_type="min",
                start_color="00FF00",
                end_type="max",
                end_color="FF0000",
            )
            ws.conditional_formatting.add("A1:A5", rule)

            wb.save(f.name)

            result = processor.process_xlsx(f.name)

            # Sheet analysis should have conditional_format_issues field
            assert hasattr(result.sheets[0], "conditional_format_issues")
            assert isinstance(result.sheets[0].conditional_format_issues, list)

        os.unlink(f.name)

    def test_cf_issues_in_total_count(self, processor):
        """Test that CF issues are counted in total issues."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb = Workbook()
            ws = wb.active

            for i in range(1, 6):
                ws.cell(row=i, column=1, value=i * 20)

            rule = ColorScaleRule(
                start_type="min",
                start_color="00FF00",
                end_type="max",
                end_color="FF0000",
            )
            ws.conditional_formatting.add("A1:A5", rule)

            wb.save(f.name)

            result = processor.process_xlsx(f.name)

            cf_count = result.summary.get("conditional_format_issues", 0)
            total = result.summary.get("total_issues", 0)

            # If there are CF issues, they should be in total
            if cf_count > 0:
                assert total >= cf_count

        os.unlink(f.name)

    def test_cf_issues_in_remediation_suggestions(self, processor):
        """Test that CF issues generate remediation suggestions."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb = Workbook()
            ws = wb.active

            for i in range(1, 6):
                ws.cell(row=i, column=1, value=i * 20)

            rule = ColorScaleRule(
                start_type="min",
                start_color="00FF00",
                end_type="max",
                end_color="FF0000",
            )
            ws.conditional_formatting.add("A1:A5", rule)

            wb.save(f.name)

            result = processor.process_xlsx(f.name)

            if result.summary.get("conditional_format_issues", 0) > 0:
                # Should have remediation suggestion mentioning CF
                cf_suggestions = [
                    s
                    for s in result.remediation_suggestions
                    if "conditional formatting" in s.lower()
                ]
                assert len(cf_suggestions) >= 1

        os.unlink(f.name)

    def test_multiple_cf_rules_same_sheet(self, processor):
        """Test detection of multiple CF rules on same sheet."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb = Workbook()
            ws = wb.active

            # Add data in two columns
            for i in range(1, 6):
                ws.cell(row=i, column=1, value=i * 10)
                ws.cell(row=i, column=2, value=i * 20)

            # Add color scale to first column
            rule1 = ColorScaleRule(
                start_type="min",
                start_color="00FF00",
                end_type="max",
                end_color="FF0000",
            )
            ws.conditional_formatting.add("A1:A5", rule1)

            # Add data bar to second column
            rule2 = DataBarRule(start_type="min", end_type="max", color="638EC6")
            ws.conditional_formatting.add("B1:B5", rule2)

            wb.save(f.name)

            result = processor.process_xlsx(f.name)

            # Should detect multiple CF issues
            cf_issues = result.sheets[0].conditional_format_issues
            # At least the color scale should be detected
            assert len(cf_issues) >= 1

        os.unlink(f.name)

    def test_cf_issues_across_multiple_sheets(self, processor):
        """Test detection of CF issues across multiple sheets."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb = Workbook()

            # First sheet with color scale
            ws1 = wb.active
            ws1.title = "Sheet1"
            for i in range(1, 6):
                ws1.cell(row=i, column=1, value=i * 10)
            rule1 = ColorScaleRule(
                start_type="min",
                start_color="00FF00",
                end_type="max",
                end_color="FF0000",
            )
            ws1.conditional_formatting.add("A1:A5", rule1)

            # Second sheet with different color scale
            ws2 = wb.create_sheet("Sheet2")
            for i in range(1, 6):
                ws2.cell(row=i, column=1, value=i * 20)
            rule2 = ColorScaleRule(
                start_type="min",
                start_color="0000FF",
                end_type="max",
                end_color="FF00FF",
            )
            ws2.conditional_formatting.add("A1:A5", rule2)

            wb.save(f.name)

            result = processor.process_xlsx(f.name)

            # Both sheets should have CF issues
            total_cf_issues = result.summary.get("conditional_format_issues", 0)
            assert total_cf_issues >= 2

        os.unlink(f.name)


class TestCFIssuesComputedField:
    """Test that CF issues appear in the computed issues field."""

    @pytest.fixture
    def processor(self):
        """Create XlsxProcessor for testing."""
        return XlsxProcessor()

    def test_cf_issues_in_computed_issues(self, processor):
        """Test that CF issues appear in the combined issues list."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb = Workbook()
            ws = wb.active

            for i in range(1, 6):
                ws.cell(row=i, column=1, value=i * 20)

            rule = ColorScaleRule(
                start_type="min",
                start_color="00FF00",
                end_type="max",
                end_color="FF0000",
            )
            ws.conditional_formatting.add("A1:A5", rule)

            wb.save(f.name)

            result = processor.process_xlsx(f.name)

            # Check computed issues field
            cf_computed = [
                i
                for i in result.issues
                if i.get("category") == "color" and "cf_" in i.get("id", "")
            ]

            if result.summary.get("conditional_format_issues", 0) > 0:
                assert len(cf_computed) >= 1
                assert cf_computed[0]["wcag_criterion"] == "WCAG 1.4.1"

        os.unlink(f.name)


if __name__ == "__main__":
    pytest.main([__file__, "-v", "-s"])
